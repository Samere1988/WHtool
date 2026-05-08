import pandas as pd
import json
import datetime
from datetime import date, timedelta
from collections import Counter
from io import BytesIO
from copy import copy
import os
from types import SimpleNamespace

from django.conf import settings
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Sum, Count, Max
from django.utils import timezone, dateparse
from django.utils.dateparse import parse_date

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    CustomerList, RunSheet, OrderArchive, FinalizedRunSheet,
    DailyRunSheetCommit, DailyRunSheetEntry, EmployeeDailyStat,
    OutboundLoad, OutboundPhoto, Vendor, PickupLog,
    PickupPhotoLog, PickupPhoto, Container, ContainerPhoto
)


# ==========================================
# --- 1. HELPER FUNCTIONS ---
# ==========================================

TRANSPORT_REGIONS = [
    "North Shore",
    "Drummond",
    "Quebec",
    "Beauce",
    "Montreal",
    "South Shore",
    "Ontario",
    "Sherbrooke",
]

TRANSPORT_REGION_BLOCKS = {
    "North Shore": {
        "region_cell": "A1",
        "driver_cell": "B1",
        "header_row": 2,
        "start_row": 3,
        "end_row": 17,
        "total_row": 18,
        "start_col": 1,
        "hidden_ids_col": 27,  # AA
    },
    "Drummond": {
        "region_cell": "K1",
        "driver_cell": "L1",
        "header_row": 2,
        "start_row": 3,
        "end_row": 17,
        "total_row": 18,
        "start_col": 11,
        "hidden_ids_col": 28,  # AB
    },
    "Quebec": {
        "region_cell": "A20",
        "driver_cell": "B20",
        "header_row": 21,
        "start_row": 22,
        "end_row": 36,
        "total_row": 37,
        "start_col": 1,
        "hidden_ids_col": 27,
    },
    "Beauce": {
        "region_cell": "K20",
        "driver_cell": "L20",
        "header_row": 21,
        "start_row": 22,
        "end_row": 36,
        "total_row": 37,
        "start_col": 11,
        "hidden_ids_col": 28,
    },
    "Montreal": {
        "region_cell": "A39",
        "driver_cell": "B39",
        "header_row": 40,
        "start_row": 41,
        "end_row": 54,
        "total_row": 55,
        "start_col": 1,
        "hidden_ids_col": 27,
    },
    "South Shore": {
        "region_cell": "K39",
        "driver_cell": "L39",
        "header_row": 40,
        "start_row": 41,
        "end_row": 54,
        "total_row": 55,
        "start_col": 11,
        "hidden_ids_col": 28,
    },
    "Ontario": {
        "region_cell": "A58",
        "driver_cell": "B58",
        "header_row": 59,
        "start_row": 60,
        "end_row": 74,
        "total_row": 75,
        "start_col": 1,
        "hidden_ids_col": 27,
    },
    "Sherbrooke": {
        "region_cell": "K58",
        "driver_cell": "L58",
        "header_row": 59,
        "start_row": 60,
        "end_row": 74,
        "total_row": 75,
        "start_col": 11,
        "hidden_ids_col": 28,
    },
}


def get_next_business_day():
    """Calculates the next shipping day, skipping weekends."""
    today = date.today()
    if today.weekday() == 4:  # Friday -> Monday
        days_to_add = 3
    elif today.weekday() == 5:  # Saturday -> Monday
        days_to_add = 2
    else:  # Sun-Thu -> Next Day
        days_to_add = 1
    return today + timedelta(days=days_to_add)


def get_selected_shipping_date(request):
    """
    Gets the selected shipping date from GET/POST.
    If not supplied, falls back to session.
    If no session date exists, defaults to next business day.
    """
    selected_date = (
        request.GET.get("shipping_date")
        or request.POST.get("shipping_date")
        or request.session.get("selected_shipping_date")
    )

    if selected_date:
        parsed = parse_date(str(selected_date))
        if parsed:
            request.session["selected_shipping_date"] = parsed.strftime("%Y-%m-%d")
            return parsed

    default_date = get_next_business_day()
    request.session["selected_shipping_date"] = default_date.strftime("%Y-%m-%d")
    return default_date


def redirect_run_sheet_for_date(shipping_date):
    return redirect(f"/run-sheet/?shipping_date={shipping_date.strftime('%Y-%m-%d')}")


def normalize_transport_region(region):
    if not region:
        return ""

    cleaned = str(region).strip()

    aliases = {
        "north shore": "North Shore",
        "drummond": "Drummond",
        "quebec": "Quebec",
        "québec": "Quebec",
        "beauce": "Beauce",
        "montreal": "Montreal",
        "montréal": "Montreal",
        "south shore": "South Shore",
        "ontario": "Ontario",
        "sherbrooke": "Sherbrooke",
    }

    return aliases.get(cleaned.lower(), cleaned)


def get_transport_group_key(order):
    """
    Controls what gets combined into one transport stop.
    Orders are combined when they share customer/vendor, city, address,
    postal code, closing time, and pickup/return type.
    """
    return (
        normalize_transport_region(order.region),
        (order.customer_name or "").strip().upper(),
        (order.city or "").strip().upper(),
        (order.address or "").strip().upper(),
        (order.postal_code or "").strip().upper(),
        (order.closing_time or "").strip().upper(),
        bool(order.is_pickup),
        bool(order.is_return),
    )


def build_transport_stops_for_region(region, shipping_date):
    """
    Converts individual RunSheet rows into combined transport stops.
    Website rows stay separate; Excel rows are combined for transport only.
    """
    orders = RunSheet.objects.filter(
        region=region,
        shipping_date=shipping_date,
    ).order_by(
        "load_index",
        "customer_name",
        "order_number",
        "id",
    )

    grouped = {}

    for order in orders:
        key = get_transport_group_key(order)

        if key not in grouped:
            grouped[key] = {
                "ids": [],
                "customer_name": order.customer_name or "",
                "city": order.city or "",
                "weight": 0,
                "skids": 0,
                "bundles": 0,
                "coils": 0,
                "closing_time": order.closing_time or "",
                "pickup": "Y" if order.is_pickup else "N",
                "is_pickup": bool(order.is_pickup),
                "is_return": bool(order.is_return),
                "driver_name": order.driver_name or "",
                "first_load_index": order.load_index or 0,
            }

        grouped[key]["ids"].append(str(order.id))
        grouped[key]["weight"] += order.weight or 0
        grouped[key]["skids"] += order.skids or 0
        grouped[key]["bundles"] += order.bundles or 0
        grouped[key]["coils"] += order.coils or 0

        if not grouped[key]["driver_name"] and order.driver_name:
            grouped[key]["driver_name"] = order.driver_name

    stops = list(grouped.values())
    stops.sort(key=lambda x: (x["first_load_index"], x["customer_name"]))
    return stops


def copy_excel_cell_style(source_cell, target_cell):
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def clear_transport_block(ws, block):
    start_col = block["start_col"]
    hidden_ids_col = block["hidden_ids_col"]

    for row in range(block["start_row"], block["end_row"] + 1):
        ws.row_dimensions[row].height = 18

        for col_offset in range(9):
            col = start_col + col_offset
            cell = ws.cell(row=row, column=col)
            source_cell = ws.cell(row=block["start_row"], column=col)

            cell.value = None
            copy_excel_cell_style(source_cell, cell)

        ws.cell(row=row, column=hidden_ids_col).value = None


def write_transport_headers(ws, block):
    start_col = block["start_col"]
    header_row = block["header_row"]

    headers = [
        "Customer Name",
        "City",
        "Weight",
        "Skids",
        "Bundles",
        "Coils",
        "Closes at",
        "Pickup",
        "",
    ]

    for idx, header in enumerate(headers):
        ws.cell(row=header_row, column=start_col + idx).value = header


def write_transport_totals(ws, block):
    start_col = block["start_col"]
    start_row = block["start_row"]
    end_row = block["end_row"]
    total_row = block["total_row"]

    weight_col = start_col + 2
    skids_col = start_col + 3
    bundles_col = start_col + 4
    coils_col = start_col + 5

    ws.cell(row=total_row, column=weight_col).value = (
        f"=SUM({ws.cell(start_row, weight_col).coordinate}:{ws.cell(end_row, weight_col).coordinate})"
    )
    ws.cell(row=total_row, column=skids_col).value = (
        f"=SUM({ws.cell(start_row, skids_col).coordinate}:{ws.cell(end_row, skids_col).coordinate})"
    )
    ws.cell(row=total_row, column=bundles_col).value = (
        f"=SUM({ws.cell(start_row, bundles_col).coordinate}:{ws.cell(end_row, bundles_col).coordinate})"
    )
    ws.cell(row=total_row, column=coils_col).value = (
        f"=SUM({ws.cell(start_row, coils_col).coordinate}:{ws.cell(end_row, coils_col).coordinate})"
    )


def apply_transport_column_widths(ws):
    """
    Auto-sizes visible transport columns based on text in the sheet,
    while keeping maximum widths reasonable.
    """
    min_widths = {
        "A": 24, "B": 16, "C": 12, "D": 9, "E": 10, "F": 9, "G": 14, "H": 10, "I": 3,
        "J": 4,
        "K": 24, "L": 16, "M": 12, "N": 9, "O": 10, "P": 9, "Q": 14, "R": 10, "S": 3,
    }

    max_widths = {
        "A": 45, "B": 28,
        "K": 45, "L": 28,
    }

    for col in range(1, 20):  # A to S
        column_letter = get_column_letter(col)
        max_length = 0

        for cell in ws[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        calculated_width = max_length + 2
        min_width = min_widths.get(column_letter, 10)
        max_width = max_widths.get(column_letter, 16)

        ws.column_dimensions[column_letter].width = min(
            max(calculated_width, min_width),
            max_width,
        )

    ws.column_dimensions["J"].width = 4
    ws.column_dimensions["AA"].hidden = True
    ws.column_dimensions["AB"].hidden = True


def calculate_performance(orders_queryset):
    """Calculates picker KPIs for the Stats page (Chart.js version)."""
    picker_counts = Counter()
    line_counts = Counter()

    for o in orders_queryset:
        if o.bar_prep:
            names = [n.strip() for n in o.bar_prep.split(',') if n.strip()]
            for name in names:
                picker_counts[name] += 1
                line_counts[name] += (o.bar_lines or 0)
        if o.sheet_prep:
            names = [n.strip() for n in o.sheet_prep.split(',') if n.strip()]
            for name in names:
                picker_counts[name] += 1
                line_counts[name] += (o.sheet_lines or 0)
        if o.covering_prep:
            names = [n.strip() for n in o.covering_prep.split(',') if n.strip()]
            for name in names:
                picker_counts[name] += 1
                line_counts[name] += (o.covering_lines or 0)

    performance = []
    labels = []
    orders_data = []
    lines_data = []

    sorted_names = sorted(picker_counts.keys(), key=lambda n: line_counts[n], reverse=True)

    for name in sorted_names:
        tot_orders = picker_counts[name]
        tot_lines = line_counts[name]
        avg = round(tot_lines / tot_orders, 1) if tot_orders > 0 else 0

        performance.append({
            'prepared_by': name,
            'total_orders': tot_orders,
            'total_lines': tot_lines,
            'avg_lines': avg,
        })
        labels.append(name)
        orders_data.append(tot_orders)
        lines_data.append(tot_lines)

    return {
        'table': performance,
        'labels': json.dumps(labels),
        'orders': json.dumps(orders_data),
        'lines': json.dumps(lines_data),
    }


def build_grouped_run_sheet_orders(orders):
    """
    Groups multiple RunSheet rows into one visible delivery stop.
    This does NOT merge anything in the database.
    """
    grouped = {}

    for order in orders:
        key = (
            order.customer_id or "",
            order.customer_name or "",
            order.city or "",
            order.address or "",
            order.postal_code or "",
            order.closing_time or "",
            bool(order.is_pickup),
            bool(order.is_return),
        )

        if key not in grouped:
            grouped[key] = {
                "main_order": order,
                "orders": [],
                "customer_name": order.customer_name or "",
                "city": order.city or "",
                "weight": 0,
                "skids": 0,
                "bundles": 0,
                "coils": 0,
                "closing_time": order.closing_time or "",
                "is_pickup": bool(order.is_pickup),
                "is_return": bool(order.is_return),
                "load_index": order.load_index or 0,
            }

        grouped[key]["orders"].append(order)
        grouped[key]["weight"] += order.weight or 0
        grouped[key]["skids"] += order.skids or 0
        grouped[key]["bundles"] += order.bundles or 0
        grouped[key]["coils"] += order.coils or 0

        if (order.load_index or 0) < grouped[key]["load_index"]:
            grouped[key]["load_index"] = order.load_index or 0
            grouped[key]["main_order"] = order

    grouped_orders = list(grouped.values())

    grouped_orders.sort(
        key=lambda stop: (
            stop["load_index"],
            stop["customer_name"],
            stop["main_order"].id,
        )
    )

    for stop in grouped_orders:
        stop["orders"].sort(
            key=lambda order: (
                order.order_number or "",
                order.id,
            )
        )

    return grouped_orders


# ==========================================
# --- 2. MAIN DASHBOARDS & STATS ---
# ==========================================
@login_required
def photos_home(request):
    return render(request, "core/photos_home.html")

@login_required
def home(request):
    return render(request, 'core/home.html')


@login_required
def run_sheet(request):
    """
    Main dispatch dashboard.
    Shows one grouped row per customer/vendor stop for the selected SHIPPING DATE.
    """
    selected_shipping_date = get_selected_shipping_date(request)

    grouped_orders = {}

    for region in TRANSPORT_REGIONS:
        orders = RunSheet.objects.filter(
            region=region,
            shipping_date=selected_shipping_date,
        ).order_by(
            "load_index",
            "customer_name",
            "order_number",
            "id",
        )

        if orders.exists():
            grouped_stops = build_grouped_run_sheet_orders(orders)

            grouped_orders[region] = {
                "orders": grouped_stops,
                "totals": {
                    "weight": sum(order.weight or 0 for order in orders),
                    "skids": sum(order.skids or 0 for order in orders),
                    "bundles": sum(order.bundles or 0 for order in orders),
                    "coils": sum(order.coils or 0 for order in orders),
                },
            }

    return render(
        request,
        "core/run_sheet.html",
        {
            "grouped_orders": grouped_orders,
            "regions": TRANSPORT_REGIONS,
            "selected_shipping_date": selected_shipping_date,
        },
    )


@login_required
def stats(request):
    """KPIs and Staff Performance Tracking."""
    now = timezone.localtime(timezone.now())
    active_tab = request.GET.get('tab', 'day')

    day_str = request.GET.get('day')
    week_str = request.GET.get('week')
    month_str = request.GET.get('month')

    target_day = parse_date(day_str) if day_str else now.date()
    day_start = timezone.make_aware(datetime.datetime.combine(target_day, datetime.time.min))
    day_end = timezone.make_aware(datetime.datetime.combine(target_day, datetime.time.max))

    if week_str and '-W' in week_str:
        year, week = int(week_str.split('-W')[0]), int(week_str.split('-W')[1])
        week_start_date = datetime.date.fromisocalendar(year, week, 1)
    else:
        week_start_date = now.date() - timedelta(days=now.weekday())
        year, week, _ = week_start_date.isocalendar()
        week_str = f"{year}-W{week:02d}"

    week_start = timezone.make_aware(datetime.datetime.combine(week_start_date, datetime.time.min))
    week_end = timezone.make_aware(datetime.datetime.combine(week_start_date + timedelta(days=6), datetime.time.max))

    if month_str and '-' in month_str:
        year, month = int(month_str.split('-')[0]), int(month_str.split('-')[1])
        month_start_date = datetime.date(year, month, 1)
    else:
        month_start_date = now.date().replace(day=1)
        month_str = month_start_date.strftime('%Y-%m')

    month_start = timezone.make_aware(datetime.datetime.combine(month_start_date, datetime.time.min))
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    month_end = next_month - timedelta(seconds=1)

    orders_day = list(OrderArchive.objects.filter(created_at__range=(day_start, day_end)))
    pickups_day = list(PickupLog.objects.filter(date_completed__range=(day_start.date(), day_end.date())))

    orders_week = list(OrderArchive.objects.filter(created_at__range=(week_start, week_end)))
    pickups_week = list(PickupLog.objects.filter(date_completed__range=(week_start.date(), week_end.date())))

    orders_month = list(OrderArchive.objects.filter(created_at__range=(month_start, month_end)))
    pickups_month = list(PickupLog.objects.filter(date_completed__range=(month_start.date(), month_end.date())))

    context = {
        'active_tab': active_tab,
        'day_val': target_day.strftime('%Y-%m-%d'),
        'week_val': week_str,
        'month_val': month_str,
        'title_day': target_day.strftime('%B %d, %Y'),
        'title_week': f"Week of {week_start_date.strftime('%B %d, %Y')}",
        'title_month': month_start_date.strftime('%B %Y'),
        'daily': calculate_performance(orders_day + pickups_day),
        'weekly': calculate_performance(orders_week + pickups_week),
        'monthly': calculate_performance(orders_month + pickups_month),
    }
    return render(request, 'core/stats.html', context)


# ==========================================
# --- 3. DISPATCH & ORDER ENTRY ---
# ==========================================

@login_required
def add_to_run_sheet(request):
    customers = CustomerList.objects.all()
    selected_shipping_date = get_selected_shipping_date(request)

    return render(
        request,
        'core/truck_dispatch/add_to_run_sheet.html',
        {
            'customers': customers,
            'selected_shipping_date': selected_shipping_date,
        },
    )


@login_required
def entry_form(request, customer_id):
    """Main manual entry form for truck orders."""
    selected_shipping_date = get_selected_shipping_date(request)
    customer = get_object_or_404(CustomerList, customer_id=str(customer_id))

    all_employees = [
        "Ahmad", "Mikey", "Michael", "Danilio", "Kasim", "Carl",
        "David", "Douglas", "Jean Duval", "Jean Thomas", "Cooper",
        "Aperam", "Elvis", "Leon", "Ismail",
    ]
    regions = ['Beauce', 'Drummond', 'Montreal', 'North Shore', 'Ontario', 'Quebec', 'Sherbrooke', 'South Shore']

    def sort_for_station(priority_names):
        top = [name for name in priority_names if name in all_employees]
        rest = sorted([name for name in all_employees if name not in priority_names])
        return top + rest

    bar_employees = sort_for_station(["Ismail", "Jean Duval", "Jean Thomas", "Aperam", "Cooper", "Elvis"])
    sheet_employees = sort_for_station(["Mikey", "Danilio", "Kasim", "Carl"])
    cov_employees = sort_for_station(["David", "Douglas", "Ahmad"])

    if request.method == "POST":
        order_num = request.POST.get('order_number', '').strip()
        if order_num and not order_num.upper().startswith('W'):
            order_num = f"W{order_num}"

        closing_time = request.POST.get('closing_time', '').strip()
        custom_address = request.POST.get('address', '').strip() or customer.address
        custom_city = request.POST.get('city', '').strip() or customer.city
        custom_postal = request.POST.get('postal_code', '').strip() or getattr(customer, 'postal_code', '')
        custom_region = request.POST.get('region', '').strip() or customer.region

        w = int(request.POST.get('weight') or 0)
        s = int(request.POST.get('skids') or 0)
        b = int(request.POST.get('bundles') or 0)
        c = int(request.POST.get('coils') or 0)

        bar_lines = int(request.POST.get('bar_lines') or 0)
        bar_pickers = request.POST.getlist('bar_prep')
        bar_prep_str = ", ".join(sorted(bar_pickers)) if bar_pickers else ""

        sheet_lines = int(request.POST.get('sheet_lines') or 0)
        sheet_pickers = request.POST.getlist('sheet_prep')
        sheet_prep_str = ", ".join(sorted(sheet_pickers)) if sheet_pickers else ""

        cov_lines = int(request.POST.get('covering_lines') or 0)
        cov_pickers = request.POST.getlist('covering_prep')
        cov_prep_str = ", ".join(sorted(cov_pickers)) if cov_pickers else ""

        total_lines = bar_lines + sheet_lines + cov_lines
        total_prep_list = list(set(bar_pickers + sheet_pickers + cov_pickers))
        total_prep_str = ", ".join(sorted(total_prep_list)) if total_prep_list else "Unknown"

        existing_customer_stop = RunSheet.objects.filter(
            customer_id=customer.customer_id,
            region=custom_region,
            city=custom_city,
            address=custom_address,
            shipping_date=selected_shipping_date,
            is_pickup=False,
            is_return=False,
        ).order_by("load_index", "id").first()

        if existing_customer_stop:
            load_index = existing_customer_stop.load_index
        else:
            max_index = RunSheet.objects.filter(
                region=custom_region,
                shipping_date=selected_shipping_date,
            ).aggregate(Max("load_index"))["load_index__max"] or 0
            load_index = max_index + 1

        OrderArchive.objects.create(
            order_number=order_num,
            customer_id=customer.customer_id,
            customer_name=customer.customer_name,
            bar_prep=bar_prep_str,
            bar_lines=bar_lines,
            sheet_prep=sheet_prep_str,
            sheet_lines=sheet_lines,
            covering_prep=cov_prep_str,
            covering_lines=cov_lines,
            prepared_by=total_prep_str,
            line_items=total_lines,
            skids=s,
            bundles=b,
            coils=c,
            weight=w,
            region=custom_region,
        )

        RunSheet.objects.create(
            customer_id=customer.customer_id,
            shipping_date=selected_shipping_date,
            customer_name=customer.customer_name,
            address=custom_address,
            city=custom_city,
            postal_code=custom_postal,
            region=custom_region,
            order_number=order_num,
            prepared_by=total_prep_str,
            line_items=total_lines,
            closing_time=closing_time,
            weight=w,
            skids=s,
            bundles=b,
            coils=c,
            load_index=load_index,
        )
        return redirect_run_sheet_for_date(selected_shipping_date)

    return render(request, 'core/truck_dispatch/entry_form.html', {
        'customer': customer,
        'bar_employees': bar_employees,
        'sheet_employees': sheet_employees,
        'cov_employees': cov_employees,
        'regions': regions,
        'selected_shipping_date': selected_shipping_date,
    })


@login_required
def edit_order(request, pk):
    """Edits granular line item details for an existing Run Sheet order."""
    run_item = get_object_or_404(RunSheet, pk=pk)
    selected_shipping_date = run_item.shipping_date

    if run_item.customer_id:
        customer = CustomerList.objects.filter(customer_id=run_item.customer_id).first()
    else:
        customer = None

    if customer is None:
        customer = SimpleNamespace(
            customer_id=run_item.customer_id or "",
            customer_name=run_item.customer_name or "",
            address=run_item.address or "",
            city=run_item.city or "",
            postal_code=run_item.postal_code or "",
            region=run_item.region or "",
        )

    archive = OrderArchive.objects.filter(order_number=run_item.order_number).order_by('-id').first()

    saved_bar = archive.bar_prep.split(', ') if archive and archive.bar_prep else []
    saved_sheet = archive.sheet_prep.split(', ') if archive and archive.sheet_prep else []
    saved_cov = archive.covering_prep.split(', ') if archive and archive.covering_prep else []

    all_employees = [
        "Ahmad", "Mikey", "Michael", "Danilio", "Kasim", "Carl",
        "David", "Douglas", "Jean Duval", "Jean Thomas", "Cooper",
        "Aperam", "Elvis", "Leon", "Ismail",
    ]

    def sort_for_station(priority_names):
        top = [name for name in priority_names if name in all_employees]
        rest = sorted([name for name in all_employees if name not in priority_names])
        return top + rest

    bar_employees = sort_for_station(["Ismail", "Jean Duval", "Jean Thomas", "Aperam", "Cooper", "Elvis"])
    sheet_employees = sort_for_station(["Mikey", "Danilio", "Kasim", "Carl"])
    cov_employees = sort_for_station(["David", "Douglas", "Ahmad"])
    regions = ['Beauce', 'Drummond', 'Montreal', 'North Shore', 'Ontario', 'Quebec', 'Sherbrooke', 'South Shore']

    if request.method == "POST":
        posted_shipping_date = get_selected_shipping_date(request)
        order_num = request.POST.get('order_number', '').strip()
        if order_num and not order_num.upper().startswith('W') and not order_num.upper().startswith('K'):
            order_num = f"W{order_num}"

        closing_time = request.POST.get('closing_time', '').strip()
        w = int(request.POST.get('weight') or 0)
        s = int(request.POST.get('skids') or 0)
        b = int(request.POST.get('bundles') or 0)
        c = int(request.POST.get('coils') or 0)

        bar_lines = int(request.POST.get('bar_lines') or 0)
        bar_pickers = request.POST.getlist('bar_prep')
        bar_prep_str = ", ".join(sorted(bar_pickers)) if bar_pickers else ""

        sheet_lines = int(request.POST.get('sheet_lines') or 0)
        sheet_pickers = request.POST.getlist('sheet_prep')
        sheet_prep_str = ", ".join(sorted(sheet_pickers)) if sheet_pickers else ""

        cov_lines = int(request.POST.get('covering_lines') or 0)
        cov_pickers = request.POST.getlist('covering_prep')
        cov_prep_str = ", ".join(sorted(cov_pickers)) if cov_pickers else ""

        total_lines = bar_lines + sheet_lines + cov_lines
        total_prep_list = list(set(bar_pickers + sheet_pickers + cov_pickers))
        total_prep_str = ", ".join(sorted(total_prep_list)) if total_prep_list else run_item.prepared_by or "Unknown"

        run_item.order_number = order_num
        run_item.address = request.POST.get('address', run_item.address)
        run_item.city = request.POST.get('city', run_item.city)
        run_item.postal_code = request.POST.get('postal_code', run_item.postal_code)
        run_item.region = request.POST.get('region', run_item.region)
        run_item.shipping_date = posted_shipping_date
        run_item.closing_time = closing_time
        run_item.weight = w
        run_item.skids = s
        run_item.bundles = b
        run_item.coils = c
        run_item.prepared_by = total_prep_str
        run_item.line_items = total_lines
        run_item.save()

        if archive:
            archive.order_number = order_num
            archive.bar_prep = bar_prep_str
            archive.bar_lines = bar_lines
            archive.sheet_prep = sheet_prep_str
            archive.sheet_lines = sheet_lines
            archive.covering_prep = cov_prep_str
            archive.covering_lines = cov_lines
            archive.prepared_by = total_prep_str
            archive.line_items = total_lines
            archive.skids = s
            archive.bundles = b
            archive.coils = c
            archive.weight = w
            archive.region = run_item.region
            archive.save()

        return redirect_run_sheet_for_date(posted_shipping_date)

    return render(request, 'core/truck_dispatch/entry_form.html', {
        'customer': customer,
        'order': run_item,
        'archive': archive,
        'saved_bar': saved_bar,
        'saved_sheet': saved_sheet,
        'saved_cov': saved_cov,
        'bar_employees': bar_employees,
        'sheet_employees': sheet_employees,
        'cov_employees': cov_employees,
        'regions': regions,
        'selected_shipping_date': selected_shipping_date,
    })


@login_required
def edit_specific_order(request, pk):
    """Simple edit form for basic order details."""
    order = get_object_or_404(RunSheet, pk=pk)
    employees = ["Ahmad", "Mikey", "Michael", "..."]
    if request.method == "POST":
        order.order_number = request.POST.get('order_number')
        order.weight = int(request.POST.get('weight') or 0)
        order.skids = int(request.POST.get('skids') or 0)
        order.save()
        return redirect_run_sheet_for_date(order.shipping_date)
    return render(request, 'core/edit_order_form.html', {'order': order, 'employees': sorted(employees)})


@login_required
def delete_stop(request, pk):
    stop = get_object_or_404(RunSheet, pk=pk)
    shipping_date = stop.shipping_date
    order_num = stop.order_number
    customer_name = stop.customer_name
    stop.delete()

    # Delete the matching archive entry only for true orders; transport-only pickups/returns do not need archive entries.
    if order_num and not (str(order_num).upper().startswith("K") or str(customer_name).upper().startswith("PICKUP:")):
        OrderArchive.objects.filter(order_number=order_num).delete()

    messages.warning(request, f"Order {order_num} for {customer_name} removed from truck and stats.")
    return redirect_run_sheet_for_date(shipping_date)


@login_required
def clear_run_sheet(request):
    selected_shipping_date = get_selected_shipping_date(request)
    if request.method == "POST":
        RunSheet.objects.filter(shipping_date=selected_shipping_date).delete()
        messages.warning(request, f"Run sheet for {selected_shipping_date} has been cleared.")
    return redirect_run_sheet_for_date(selected_shipping_date)


@login_required
def commit_and_clear_day(request):
    """
    Finalizes the selected shipping date.
    Saves every RunSheet row to DailyRunSheetEntry, employee stats to EmployeeDailyStat,
    then clears only that selected shipping date from the live board.
    """
    if request.method != "POST":
        return redirect("run_sheet")

    shipping_date = get_selected_shipping_date(request)

    current_stops = RunSheet.objects.filter(
        shipping_date=shipping_date,
    ).order_by(
        "region",
        "load_index",
        "customer_name",
        "order_number",
        "id",
    )

    if not current_stops.exists():
        messages.info(request, "Run sheet is already empty. Nothing to commit.")
        return redirect_run_sheet_for_date(shipping_date)

    with transaction.atomic():
        commit = DailyRunSheetCommit.objects.create(
            shipping_date=shipping_date,
            total_weight=sum(stop.weight or 0 for stop in current_stops),
            total_skids=sum(stop.skids or 0 for stop in current_stops),
            total_bundles=sum(stop.bundles or 0 for stop in current_stops),
            total_coils=sum(stop.coils or 0 for stop in current_stops),
        )

        for stop in current_stops:
            DailyRunSheetEntry.objects.create(
                commit=commit,
                original_run_sheet_id=stop.id,
                customer_id=stop.customer_id,
                customer_name=stop.customer_name,
                order_number=stop.order_number,
                address=stop.address,
                city=stop.city,
                postal_code=stop.postal_code,
                region=stop.region,
                driver_name=stop.driver_name,
                load_index=stop.load_index,
                closing_time=stop.closing_time,
                weight=stop.weight or 0,
                skids=stop.skids or 0,
                bundles=stop.bundles or 0,
                coils=stop.coils or 0,
                is_pickup=stop.is_pickup,
                is_return=stop.is_return,
                prepared_by=stop.prepared_by,
                line_items=stop.line_items or 0,
            )

        order_numbers = [stop.order_number for stop in current_stops if stop.order_number]
        archives = OrderArchive.objects.filter(order_number__in=order_numbers)

        employee_stats = {}

        def get_employee_row(employee_name):
            employee_name = employee_name.strip()
            if employee_name not in employee_stats:
                employee_stats[employee_name] = {
                    "orders_picked": 0,
                    "total_lines": 0,
                    "bar_orders": 0,
                    "bar_lines": 0,
                    "sheet_orders": 0,
                    "sheet_lines": 0,
                    "covering_orders": 0,
                    "covering_lines": 0,
                }
            return employee_stats[employee_name]

        def split_names(value):
            if not value:
                return []
            return [name.strip() for name in str(value).split(",") if name.strip()]

        for archive in archives:
            for name in split_names(archive.bar_prep):
                row = get_employee_row(name)
                row["orders_picked"] += 1
                row["total_lines"] += archive.bar_lines or 0
                row["bar_orders"] += 1
                row["bar_lines"] += archive.bar_lines or 0

            for name in split_names(archive.sheet_prep):
                row = get_employee_row(name)
                row["orders_picked"] += 1
                row["total_lines"] += archive.sheet_lines or 0
                row["sheet_orders"] += 1
                row["sheet_lines"] += archive.sheet_lines or 0

            for name in split_names(archive.covering_prep):
                row = get_employee_row(name)
                row["orders_picked"] += 1
                row["total_lines"] += archive.covering_lines or 0
                row["covering_orders"] += 1
                row["covering_lines"] += archive.covering_lines or 0

        for employee_name, stats in employee_stats.items():
            EmployeeDailyStat.objects.create(
                commit=commit,
                employee_name=employee_name,
                orders_picked=stats["orders_picked"],
                total_lines=stats["total_lines"],
                bar_orders=stats["bar_orders"],
                bar_lines=stats["bar_lines"],
                sheet_orders=stats["sheet_orders"],
                sheet_lines=stats["sheet_lines"],
                covering_orders=stats["covering_orders"],
                covering_lines=stats["covering_lines"],
            )

        RunSheet.objects.filter(shipping_date=shipping_date).delete()

    messages.success(
        request,
        f"Run sheet committed for {shipping_date.strftime('%A %B')} {shipping_date.day}. "
        "The selected shipping date has been cleared.",
    )
    return redirect_run_sheet_for_date(shipping_date)


@login_required
def finalize_run_sheet(request):
    shipping_date = get_selected_shipping_date(request)
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    grouped_orders = {}

    for region in TRANSPORT_REGIONS:
        orders = RunSheet.objects.filter(
            region=region,
            shipping_date=shipping_date,
        ).order_by(
            "load_index",
            "customer_name",
            "order_number",
            "id",
        )
        if orders.exists():
            cust_ids = orders.values_list('customer_id', flat=True)
            detailed_archive = OrderArchive.objects.filter(
                customer_id__in=cust_ids,
                created_at__gte=today_start,
            ).order_by('created_at')
            grouped_orders[region] = {
                'orders': orders,
                'detailed_archive': detailed_archive,
                'totals': {
                    'weight': sum(o.weight or 0 for o in orders),
                    'skids': sum(o.skids or 0 for o in orders),
                    'bundles': sum(o.bundles or 0 for o in orders),
                    'coils': sum(o.coils or 0 for o in orders),
                },
            }
    return render(request, 'core/finalize_view.html', {
        'grouped_orders': grouped_orders,
        'shipping_date': shipping_date,
    })


@login_required
def run_sheet_history(request):
    history = FinalizedRunSheet.objects.all().order_by('-finalized_at')
    return render(request, 'core/history.html', {'history': history})


# ==========================================
# --- 4. EXCEL IMPORT / EXPORT ---
# ==========================================

@login_required
def upload_run_sheet(request):
    """
    Imports a returned transport Excel sheet for the selected shipping date.
    It updates region/driver/load order using hidden IDs and deletes only entries
    from that selected shipping date that were not included in the import.
    """
    selected_shipping_date = get_selected_shipping_date(request)

    if request.method != "POST" or not request.FILES.get("excel_file"):
        return redirect_run_sheet_for_date(selected_shipping_date)

    file = request.FILES["excel_file"]

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb["Run Sheet"]

        imported_ids = set()
        missing_ids = []
        updated_count = 0

        for region in TRANSPORT_REGIONS:
            block = TRANSPORT_REGION_BLOCKS[region]

            driver_text = ws[block["driver_cell"]].value or ""
            driver_name = ""

            if "driver:" in str(driver_text).lower():
                driver_name = str(driver_text).split(":", 1)[1].strip()

            load_index = 1

            for row in range(block["start_row"], block["end_row"] + 1):
                hidden_ids_value = ws.cell(row=row, column=block["hidden_ids_col"]).value

                if not hidden_ids_value:
                    continue

                id_list = [
                    item.strip()
                    for item in str(hidden_ids_value).split(",")
                    if item.strip().isdigit()
                ]

                if not id_list:
                    continue

                for run_sheet_id in id_list:
                    imported_ids.add(int(run_sheet_id))

                    try:
                        stop = RunSheet.objects.get(id=int(run_sheet_id))
                    except RunSheet.DoesNotExist:
                        missing_ids.append(run_sheet_id)
                        continue

                    stop.region = region
                    stop.driver_name = driver_name
                    stop.load_index = load_index
                    stop.shipping_date = selected_shipping_date
                    stop.save(update_fields=["region", "driver_name", "load_index", "shipping_date"])
                    updated_count += 1

                load_index += 1

        if updated_count == 0:
            messages.error(
                request,
                "No hidden RunSheet IDs were found. Make sure you are importing a run sheet exported from this website.",
            )
            return redirect_run_sheet_for_date(selected_shipping_date)

        deleted_count, _ = RunSheet.objects.filter(
            shipping_date=selected_shipping_date,
        ).exclude(id__in=imported_ids).delete()

        if missing_ids:
            messages.warning(
                request,
                f"Imported {updated_count} orders and removed {deleted_count} old entries. "
                f"Some IDs from the Excel file no longer exist: {', '.join(missing_ids)}",
            )
        else:
            messages.success(
                request,
                f"Successfully imported transport run sheet. "
                f"Updated {updated_count} orders and removed {deleted_count} old entries.",
            )

    except Exception as e:
        messages.error(request, f"Upload error: {e}")

    return redirect_run_sheet_for_date(selected_shipping_date)


@login_required
def export_run_sheet_excel(request):
    selected_shipping_date = get_selected_shipping_date(request)

    template_path = os.path.join(
        settings.BASE_DIR,
        "core",
        "excel_templates",
        "transport_run_sheet_template.xlsx",
    )

    if not os.path.exists(template_path):
        messages.error(
            request,
            "Excel template not found. Add transport_run_sheet_template.xlsx to core/excel_templates/",
        )
        return redirect_run_sheet_for_date(selected_shipping_date)

    wb = load_workbook(template_path)
    ws = wb["Run Sheet"]

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    red_font = Font(color="FF0000", bold=True)

    ws.column_dimensions["AA"].hidden = True
    ws.column_dimensions["AB"].hidden = True

    for region in TRANSPORT_REGIONS:
        block = TRANSPORT_REGION_BLOCKS[region]
        stops = build_transport_stops_for_region(region, selected_shipping_date)
        max_rows = block["end_row"] - block["start_row"] + 1

        if len(stops) > max_rows:
            messages.error(
                request,
                f"{region} has {len(stops)} combined transport stops, "
                f"but the Excel template only has room for {max_rows}.",
            )
            return redirect_run_sheet_for_date(selected_shipping_date)

        clear_transport_block(ws, block)
        write_transport_headers(ws, block)

        ws[block["region_cell"]] = f"region {region}"

        driver_name = ""
        for stop in stops:
            if stop["driver_name"]:
                driver_name = stop["driver_name"]
                break

        ws[block["driver_cell"]] = f"Driver: {driver_name}" if driver_name else "Driver:"

        row = block["start_row"]
        start_col = block["start_col"]
        hidden_ids_col = block["hidden_ids_col"]

        for stop in stops:
            ws.row_dimensions[row].height = 18

            visible_values = [
                stop["customer_name"],
                stop["city"],
                stop["weight"],
                stop["skids"],
                stop["bundles"],
                stop["coils"],
                stop["closing_time"],
                stop["pickup"],
                "",
            ]

            for col_offset, value in enumerate(visible_values):
                cell = ws.cell(row=row, column=start_col + col_offset)
                cell.value = value
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=False,
                )

                if stop["is_pickup"] or stop["is_return"]:
                    cell.fill = yellow_fill
                    cell.font = red_font

            ws.cell(row=row, column=hidden_ids_col).value = ",".join(stop["ids"])
            row += 1

        write_transport_totals(ws, block)

    filename = f"{selected_shipping_date.strftime('%A %B')} {selected_shipping_date.day} runsheet.xlsx"
    apply_transport_column_widths(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ==========================================
# --- 5. CUSTOMER & VENDOR DATABASES ---
# ==========================================

@login_required
def customer_list(request):
    """Alternate view for customer list."""
    customers = CustomerList.objects.all()
    return render(request, 'core/database/customers.html', {'customers': customers})


@login_required
def manage_customers(request):
    customers = CustomerList.objects.all().order_by('customer_name')
    return render(request, 'core/database/manage_customers.html', {'customers': customers})


@login_required
def edit_customer(request, pk=None):
    regions = ["North Shore", "Drummond", "Quebec", "Beauce", "Montreal", "South Shore", "Ontario", "Sherbrooke"]
    customer = get_object_or_404(CustomerList, pk=pk) if pk else None
    title = f"Edit {customer.customer_name}" if pk else "Add New Customer"

    if request.method == "POST":
        cid = request.POST.get('customer_id')
        name = request.POST.get('customer_name')
        addr = request.POST.get('address')
        city = request.POST.get('city')
        reg = request.POST.get('region')
        post = request.POST.get('postal_code')

        if pk:
            customer.customer_id = cid
            customer.customer_name = name
            customer.address = addr
            customer.city = city
            customer.region = reg
            customer.postal_code = post
            customer.save()
        else:
            CustomerList.objects.create(
                customer_id=cid,
                customer_name=name,
                address=addr,
                city=city,
                region=reg,
                postal_code=post,
            )
        return redirect('manage_customers')

    return render(request, 'core/database/edit_customer.html', {
        'customer': customer,
        'title': title,
        'regions': sorted(regions),
    })


@login_required
def manage_vendors(request):
    vendors = Vendor.objects.all().order_by('name')
    return render(request, 'core/database/manage_vendors.html', {'vendors': vendors})


@login_required
def edit_vendor(request, pk=None):
    regions = ["North Shore", "Drummond", "Quebec", "Beauce", "Montreal", "South Shore", "Ontario", "Sherbrooke"]
    vendor = get_object_or_404(Vendor, pk=pk) if pk else None
    title = f"Edit {vendor.name}" if pk else "Add New Vendor"

    if request.method == "POST":
        name = request.POST.get('name')
        addr = request.POST.get('address')
        city = request.POST.get('city')
        reg = request.POST.get('region')
        post = request.POST.get('postal_code')

        if pk:
            vendor.name = name
            vendor.address = addr
            vendor.city = city
            vendor.region = reg
            vendor.postal_code = post
            vendor.save()
            messages.success(request, f"Vendor {name} updated.")
        else:
            Vendor.objects.create(
                name=name,
                address=addr,
                city=city,
                region=reg,
                postal_code=post,
            )
            messages.success(request, f"New vendor {name} added.")
        return redirect('manage_vendors')

    return render(request, 'core/database/edit_vendor.html', {
        'vendor': vendor,
        'title': title,
        'regions': sorted(regions),
    })


# ==========================================
# --- 6. TRUCK LOGIC: PICKUPS & RETURNS ---
# ==========================================

@login_required
def select_vendor_pickup(request):
    vendors = Vendor.objects.all().order_by("name")
    selected_shipping_date = get_selected_shipping_date(request)
    return render(request, "core/truck_dispatch/select_vendor_pickup.html", {
        "vendors": vendors,
        "selected_shipping_date": selected_shipping_date,
    })


@login_required
def add_pickup(request):
    """Legacy manual vendor pickup entry page."""
    selected_shipping_date = get_selected_shipping_date(request)
    vendors = Vendor.objects.all().order_by('name')
    regions = TRANSPORT_REGIONS

    if request.method == "POST":
        v_name = request.POST.get('vendor_name', '').strip()
        pos = request.POST.get('po_numbers', '').strip().upper()
        addr = request.POST.get('address', '').strip()
        city = request.POST.get('city', '').strip()
        region = request.POST.get('region')
        closing_time = request.POST.get('closing_time', '').strip()

        if request.POST.get('save_vendor'):
            Vendor.objects.get_or_create(name=v_name, address=addr, city=city, region=region)

        max_idx = RunSheet.objects.filter(
            region=region,
            shipping_date=selected_shipping_date,
        ).aggregate(Max('load_index'))['load_index__max'] or 0

        RunSheet.objects.create(
            shipping_date=selected_shipping_date,
            customer_name=f"PICKUP: {v_name}",
            address=addr,
            city=city,
            region=region,
            order_number=pos,
            closing_time=closing_time,
            is_pickup=True,
            load_index=max_idx + 1,
        )
        return redirect_run_sheet_for_date(selected_shipping_date)

    return render(request, 'core/truck_dispatch/add_pickup.html', {
        'vendors': vendors,
        'regions': regions,
        'selected_shipping_date': selected_shipping_date,
    })


@login_required
def add_pickup_form(request, vendor_id):
    """Adds a vendor pickup to the selected shipping date."""
    selected_shipping_date = get_selected_shipping_date(request)
    vendor = get_object_or_404(Vendor, pk=vendor_id)
    regions = TRANSPORT_REGIONS

    if request.method == "POST":
        po_numbers = request.POST.get("po_numbers", "").strip().upper()
        pickup_address = request.POST.get("address", "").strip() or vendor.address
        pickup_city = request.POST.get("city", "").strip() or vendor.city
        pickup_postal = request.POST.get("postal_code", "").strip().upper() or getattr(vendor, "postal_code", "")
        selected_region = request.POST.get("region", "").strip() or vendor.region
        closing_time = request.POST.get("closing_time", "").strip()

        weight = max(0, int(request.POST.get("weight") or 0))
        skids = max(0, int(request.POST.get("skids") or 0))
        bundles = max(0, int(request.POST.get("bundles") or 0))
        coils = max(0, int(request.POST.get("coils") or 0))

        is_redirect = request.POST.get("is_redirect") == "on"
        dest_name = request.POST.get("dest_name", "").strip()
        dest_city = request.POST.get("dest_city", "").strip()
        dest_postal = request.POST.get("dest_postal", "").strip().upper()

        if is_redirect and dest_name:
            display_name = (
                f"PICKUP: {vendor.name} ({pickup_city}) "
                f"➔ DELIVER: {dest_name} ({dest_city} {dest_postal})"
            )
        else:
            display_name = f"PICKUP: {vendor.name}"

        existing_vendor_stop = RunSheet.objects.filter(
            customer_name=display_name,
            region=selected_region,
            city=pickup_city,
            address=pickup_address,
            postal_code=pickup_postal,
            shipping_date=selected_shipping_date,
            is_pickup=True,
        ).order_by("load_index", "id").first()

        if existing_vendor_stop:
            load_index = existing_vendor_stop.load_index
        else:
            max_index = RunSheet.objects.filter(
                region=selected_region,
                shipping_date=selected_shipping_date,
            ).aggregate(Max("load_index"))["load_index__max"] or 0
            load_index = max_index + 1

        RunSheet.objects.create(
            shipping_date=selected_shipping_date,
            customer_id="",
            customer_name=display_name,
            address=pickup_address,
            city=pickup_city,
            postal_code=pickup_postal,
            region=selected_region,
            order_number=po_numbers,
            prepared_by="Vendor Pickup",
            line_items=0,
            closing_time=closing_time,
            weight=weight,
            skids=skids,
            bundles=bundles,
            coils=coils,
            is_pickup=True,
            load_index=load_index,
        )

        messages.success(request, f"Pickup for {vendor.name} added to {selected_region}.")
        return redirect_run_sheet_for_date(selected_shipping_date)

    return render(request, "core/truck_dispatch/add_pickup_form.html", {
        "vendor": vendor,
        "regions": regions,
        "selected_shipping_date": selected_shipping_date,
    })


@login_required
def select_customer_return(request):
    customers = CustomerList.objects.all().order_by("customer_name")
    selected_shipping_date = get_selected_shipping_date(request)
    return render(request, "core/truck_dispatch/select_customer_return.html", {
        "customers": customers,
        "selected_shipping_date": selected_shipping_date,
    })


@login_required
def add_return(request):
    """Legacy return entry page."""
    selected_shipping_date = get_selected_shipping_date(request)
    customers = CustomerList.objects.all().order_by('customer_name')

    if request.method == "POST":
        customer = get_object_or_404(CustomerList, customer_id=request.POST.get('customer_id'))
        ret_num = request.POST.get('return_number', '').strip().upper()
        if ret_num and not ret_num.startswith('K'):
            ret_num = f"K{ret_num}"

        max_idx = RunSheet.objects.filter(
            region=customer.region,
            shipping_date=selected_shipping_date,
        ).aggregate(Max('load_index'))['load_index__max'] or 0

        RunSheet.objects.create(
            shipping_date=selected_shipping_date,
            customer_id=customer.customer_id,
            customer_name=customer.customer_name,
            address=customer.address,
            city=customer.city,
            postal_code=customer.postal_code,
            region=customer.region,
            order_number=ret_num,
            is_return=True,
            load_index=max_idx + 1,
            prepared_by="Return Pickup",
        )
        messages.success(request, f"Return {ret_num} added to board.")
        return redirect_run_sheet_for_date(selected_shipping_date)

    return render(request, 'core/truck_dispatch/add_return.html', {
        'customers': customers,
        'selected_shipping_date': selected_shipping_date,
    })


@login_required
def add_return_form(request, customer_id):
    """Adds a customer return pickup to the selected shipping date."""
    selected_shipping_date = get_selected_shipping_date(request)
    customer = get_object_or_404(CustomerList, customer_id=str(customer_id))
    regions = ['Beauce', 'Drummond', 'Montreal', 'North Shore', 'Ontario', 'Quebec', 'Sherbrooke', 'South Shore']

    if request.method == "POST":
        return_num = request.POST.get("return_number", "").strip()
        if return_num and not return_num.upper().startswith("K"):
            return_num = f"K{return_num}"

        custom_address = request.POST.get("address", "").strip() or customer.address
        custom_city = request.POST.get("city", "").strip() or customer.city
        custom_postal = request.POST.get("postal_code", "").strip() or getattr(customer, "postal_code", "")
        custom_region = request.POST.get("region", "").strip() or customer.region
        closing_time = request.POST.get("closing_time", "").strip()

        weight = int(request.POST.get("weight") or 0)
        skids = int(request.POST.get("skids") or 0)
        bundles = int(request.POST.get("bundles") or 0)
        coils = int(request.POST.get("coils") or 0)

        existing_customer_stop = RunSheet.objects.filter(
            customer_id=customer.customer_id,
            region=custom_region,
            city=custom_city,
            address=custom_address,
            shipping_date=selected_shipping_date,
            is_return=True,
        ).order_by("load_index", "id").first()

        if existing_customer_stop:
            load_index = existing_customer_stop.load_index
        else:
            max_index = RunSheet.objects.filter(
                region=custom_region,
                shipping_date=selected_shipping_date,
            ).aggregate(Max("load_index"))["load_index__max"] or 0
            load_index = max_index + 1

        RunSheet.objects.create(
            shipping_date=selected_shipping_date,
            customer_id=customer.customer_id,
            customer_name=customer.customer_name,
            address=custom_address,
            city=custom_city,
            postal_code=custom_postal,
            region=custom_region,
            order_number=return_num,
            prepared_by="Return Pickup",
            line_items=0,
            closing_time=closing_time,
            weight=weight,
            skids=skids,
            bundles=bundles,
            coils=coils,
            is_return=True,
            load_index=load_index,
        )

        messages.success(request, f"Return pickup for {customer.customer_name} added to {custom_region}.")
        return redirect_run_sheet_for_date(selected_shipping_date)

    return render(request, "core/truck_dispatch/add_return_form.html", {
        "customer": customer,
        "regions": regions,
        "selected_shipping_date": selected_shipping_date,
    })


# ==========================================
# --- 7. COUNTER PICKUPS (PERMANENT LOG) ---
# ==========================================

@login_required
def select_customer_pickup(request):
    customers = CustomerList.objects.all().order_by('customer_name')
    return render(request, 'core/customer_pickup/select_customer_pickup.html', {'customers': customers})


@login_required
def pickup_log_list(request):
    dates = PickupLog.objects.dates('date_completed', 'day', order='DESC')
    return render(request, 'core/customer_pickup/cust_pickup.html', {'dates': dates})


@login_required
def daily_pickup_detail(request, date):
    orders = PickupLog.objects.filter(date_completed=date).order_by('customer_name', 'order_number')
    return render(request, 'core/customer_pickup/daily_pickup_detail.html', {'orders': orders, 'date': date})


@login_required
def add_pickup_order(request, customer_id):
    customer = get_object_or_404(CustomerList, customer_id=str(customer_id))
    all_employees = [
        "Ahmad", "Mikey", "Michael", "Danilio", "Kasim", "Carl",
        "David", "Douglas", "Jean Duval", "Jean Thomas", "Cooper",
        "Aperam", "Elvis", "Leon", "Ismail",
    ]

    def sort_for_station(priority_names):
        top = [name for name in priority_names if name in all_employees]
        rest = sorted([name for name in all_employees if name not in priority_names])
        return top + rest

    bar_employees = sort_for_station(["Ismail", "Jean Duval", "Jean Thomas", "Aperam", "Cooper", "Elvis"])
    sheet_employees = sort_for_station(["Mikey", "Danilio", "Kasim", "Carl"])
    cov_employees = sort_for_station(["David", "Douglas", "Ahmad"])

    if request.method == "POST":
        order_num = request.POST.get('order_number', '').strip()
        if order_num and not order_num.upper().startswith('W'):
            order_num = f"W{order_num}"

        w = max(0, int(request.POST.get('weight') or 0))
        s = max(0, int(request.POST.get('skids') or 0))
        b = max(0, int(request.POST.get('bundles') or 0))
        c = max(0, int(request.POST.get('coils') or 0))

        bar_lines = max(0, int(request.POST.get('bar_lines') or 0))
        sheet_lines = max(0, int(request.POST.get('sheet_lines') or 0))
        cov_lines = max(0, int(request.POST.get('covering_lines') or 0))

        bar_pickers = request.POST.getlist('bar_prep')
        sheet_pickers = request.POST.getlist('sheet_prep')
        cov_pickers = request.POST.getlist('covering_prep')

        PickupLog.objects.create(
            customer_name=customer.customer_name,
            customer_id=customer.customer_id,
            order_number=order_num,
            weight=w,
            skids=s,
            bundles=b,
            coils=c,
            bar_lines=bar_lines,
            sheet_lines=sheet_lines,
            covering_lines=cov_lines,
            bar_prep=", ".join(sorted(bar_pickers)) if bar_pickers else "",
            sheet_prep=", ".join(sorted(sheet_pickers)) if sheet_pickers else "",
            covering_prep=", ".join(sorted(cov_pickers)) if cov_pickers else "",
        )
        messages.success(request, f"Counter Pickup {order_num} logged.")
        return redirect('pickup_log_list')

    return render(request, 'core/truck_dispatch/entry_form.html', {
        'customer': customer,
        'bar_employees': bar_employees,
        'sheet_employees': sheet_employees,
        'cov_employees': cov_employees,
        'is_counter_pickup': True,
    })


@login_required
def edit_pickup_order(request, pk):
    order = get_object_or_404(PickupLog, pk=pk)
    all_employees = [
        "Ahmad", "Mikey", "Michael", "Danilio", "Kasim", "Carl",
        "David", "Douglas", "Jean Duval", "Jean Thomas", "Cooper",
        "Aperam", "Elvis", "Leon", "Ismail",
    ]
    saved_bar = [n.strip() for n in order.bar_prep.split(',')]
    saved_sheet = [n.strip() for n in order.sheet_prep.split(',')]
    saved_cov = [n.strip() for n in order.covering_prep.split(',')]

    if request.method == "POST":
        order.order_number = request.POST.get('order_number')
        order.weight = max(0, int(request.POST.get('weight') or 0))
        order.skids = max(0, int(request.POST.get('skids') or 0))
        order.bundles = max(0, int(request.POST.get('bundles') or 0))
        order.coils = max(0, int(request.POST.get('coils') or 0))
        order.bar_lines = max(0, int(request.POST.get('bar_lines') or 0))
        order.sheet_lines = max(0, int(request.POST.get('sheet_lines') or 0))
        order.covering_lines = max(0, int(request.POST.get('covering_lines') or 0))
        order.bar_prep = ", ".join(request.POST.getlist('bar_prep'))
        order.sheet_prep = ", ".join(request.POST.getlist('sheet_prep'))
        order.covering_prep = ", ".join(request.POST.getlist('covering_prep'))
        order.save()
        messages.warning(request, f"Pickup {order.order_number} updated.")
        return redirect('daily_pickup_detail', date=order.date_completed.strftime('%Y-%m-%d'))

    return render(request, 'core/truck_dispatch/entry_form.html', {
        'order': order,
        'is_counter_pickup': True,
        'bar_employees': all_employees,
        'sheet_employees': all_employees,
        'cov_employees': all_employees,
        'saved_bar': saved_bar,
        'saved_sheet': saved_sheet,
        'saved_cov': saved_cov,
    })


@login_required
def delete_pickup_order(request, pk):
    order = get_object_or_404(PickupLog, pk=pk)
    target_date = order.date_completed.strftime('%Y-%m-%d')
    order.delete()
    messages.error(request, "Pickup record deleted.")
    return redirect('daily_pickup_detail', date=target_date)


# ==========================================
# --- 8. PHOTO LOGS: CONTAINERS ---
# ==========================================

@login_required
def container_list(request):
    containers = Container.objects.all().order_by('-date_received')
    return render(request, 'core/containers/list.html', {'containers': containers})


@login_required
def container_detail(request, pk):
    container = get_object_or_404(Container, pk=pk)
    return render(request, 'core/containers/detail.html', {'container': container})


@login_required
def add_container(request):
    prefill_num = request.GET.get('container_num', '')

    if request.method == 'POST':
        container_num = request.POST.get('container_number', '').strip().upper()
        unloaded_by = request.POST.get('unloaded_by', '').strip()
        date_received_str = request.POST.get('date_received', '').strip()
        images = request.FILES.getlist('photos')

        selected_date = parse_date(date_received_str) if date_received_str else timezone.now().date()

        if container_num and images:
            container, created = Container.objects.get_or_create(
                container_number=container_num,
                date_received=selected_date,
                defaults={"unloaded_by": unloaded_by}
            )

            if unloaded_by and container.unloaded_by != unloaded_by:
                container.unloaded_by = unloaded_by
                container.save(update_fields=["unloaded_by"])

            for image in images:
                ContainerPhoto.objects.create(container=container, image=image)

            messages.success(request, f"Successfully uploaded {len(images)} photos.")
            return redirect('container_detail', pk=container.pk)

        messages.error(request, "Provide a container number and photo.")

    return render(
        request,
        'core/containers/add.html',
        {
            'prefill_num': prefill_num,
            "today": timezone.now().date(),
        }
    )


@login_required
def upload_more_container_photos(request, pk):
    container = get_object_or_404(Container, pk=pk)
    if request.method == 'POST':
        images = request.FILES.getlist('photos')
        if images:
            for img in images:
                ContainerPhoto.objects.create(container=container, image=img)
            messages.success(request, f"Added {len(images)} photos.")
    return redirect('container_detail', pk=container.pk)


@login_required
def delete_container(request, pk):
    container = get_object_or_404(Container, pk=pk)
    container.delete()
    messages.warning(request, "Container and photos deleted.")
    return redirect('container_list')


@login_required
def delete_container_photo(request, photo_id):
    photo = get_object_or_404(ContainerPhoto, pk=photo_id)
    container_id = photo.container.pk
    photo.delete()
    return redirect('container_detail', pk=container_id)


# ==========================================
# --- 9. PHOTO LOGS: OUTBOUND TRUCKS ---
# ==========================================

@login_required
def outbound_list(request):
    loads = OutboundLoad.objects.all().order_by('-date_loaded', 'truck_name')
    return render(request, 'core/outbound/list.html', {'loads': loads})


@login_required
def outbound_detail(request, pk):
    load = get_object_or_404(OutboundLoad, pk=pk)
    return render(request, 'core/outbound/detail.html', {'load': load})


@login_required
def add_outbound_photos(request):
    if request.method == 'POST':
        truck_name = request.POST.get('truck_name', '').strip().upper()
        loaded_by = request.POST.get('loaded_by', '').strip()
        date_loaded_str = request.POST.get('date_loaded', '').strip()
        images = request.FILES.getlist('photos')

        selected_date = parse_date(date_loaded_str) if date_loaded_str else timezone.now().date()

        if truck_name and images:
            load, created = OutboundLoad.objects.get_or_create(
                truck_name=truck_name,
                date_loaded=selected_date,
                defaults={"loaded_by": loaded_by}
            )

            if loaded_by and load.loaded_by != loaded_by:
                load.loaded_by = loaded_by
                load.save(update_fields=["loaded_by"])

            for img in images:
                OutboundPhoto.objects.create(load=load, image=img)

            messages.success(request, f"Photos uploaded for {truck_name}")
            return redirect('outbound_detail', pk=load.pk)

        messages.error(request, "Provide a truck name and photo.")

    return render(
        request,
        'core/outbound/add.html',
        {
            "today": timezone.now().date(),
        }
    )


@login_required
def upload_more_outbound_photos(request, pk):
    load = get_object_or_404(OutboundLoad, pk=pk)
    if request.method == 'POST':
        images = request.FILES.getlist('photos')
        if images:
            for img in images:
                OutboundPhoto.objects.create(load=load, image=img)
            messages.success(request, f"Added {len(images)} photos.")
    return redirect('outbound_detail', pk=load.pk)


@login_required
def delete_outbound_load(request, pk):
    load = get_object_or_404(OutboundLoad, pk=pk)
    load.delete()
    messages.warning(request, "Load and photos deleted.")
    return redirect('outbound_list')


@login_required
def delete_outbound_photo(request, photo_id):
    photo = get_object_or_404(OutboundPhoto, pk=photo_id)
    load_id = photo.load.pk
    photo.delete()
    return redirect('outbound_detail', pk=load_id)


# ==========================================
# --- 10. PHOTO LOGS: CUSTOMER PICKUPS ---
# ==========================================

@login_required
def pickup_photo_list(request):
    logs = PickupPhotoLog.objects.all().order_by('-date_picked_up', 'customer_name')
    return render(request, 'core/pickups/list.html', {'logs': logs})


@login_required
def pickup_photo_detail(request, pk):
    log = get_object_or_404(PickupPhotoLog, pk=pk)
    return render(request, 'core/pickups/detail.html', {'log': log})


@login_required
def add_pickup_photos(request):
    if request.method == 'POST':
        cust_name = request.POST.get('customer_name', '').strip().upper()
        order_num = request.POST.get('order_number', '').strip().upper()
        loaded_by = request.POST.get('loaded_by', '').strip()
        date_picked_up_str = request.POST.get('date_picked_up', '').strip()
        images = request.FILES.getlist('photos')

        selected_date = parse_date(date_picked_up_str) if date_picked_up_str else timezone.now().date()

        if cust_name and images:
            log, created = PickupPhotoLog.objects.get_or_create(
                customer_name=cust_name,
                order_number=order_num,
                date_picked_up=selected_date,
                defaults={"loaded_by": loaded_by}
            )

            if loaded_by and log.loaded_by != loaded_by:
                log.loaded_by = loaded_by
                log.save(update_fields=["loaded_by"])

            for img in images:
                PickupPhoto.objects.create(log=log, image=img)

            messages.success(request, f"Photos saved for {cust_name}")
            return redirect('pickup_photo_detail', pk=log.pk)

        messages.error(request, "Provide a customer name and photo.")

    return render(
        request,
        'core/pickups/add.html',
        {
            "today": timezone.now().date(),
        }
    )


@login_required
def upload_more_pickup_photos(request, pk):
    log = get_object_or_404(PickupPhotoLog, pk=pk)
    if request.method == 'POST':
        images = request.FILES.getlist('photos')
        if images:
            for img in images:
                PickupPhoto.objects.create(log=log, image=img)
            messages.success(request, f"Added {len(images)} photos.")
    return redirect('pickup_photo_detail', pk=log.pk)


@login_required
def delete_pickup_photo_log(request, pk):
    log = get_object_or_404(PickupPhotoLog, pk=pk)
    log.delete()
    messages.warning(request, "Pickup photo folder deleted.")
    return redirect('pickup_photo_list')


@login_required
def delete_pickup_individual_photo(request, photo_id):
    photo = get_object_or_404(PickupPhoto, pk=photo_id)
    log_id = photo.log.pk
    photo.delete()
    return redirect('pickup_photo_detail', pk=log_id)
