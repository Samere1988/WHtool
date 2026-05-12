from copy import copy
from io import BytesIO

from django.contrib.auth.decorators import login_required

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .models import RunSheet
from . import views as base_views


def _metadata_col_for_block(block):
    # Existing layout uses AA/AB hidden columns under hidden_ids_col.
    # Keep this only to clear old hidden metadata so stale IDs cannot remain behind.
    return block.get("customer_code_col") or block.get("hidden_ids_col")


def _customer_code_for_stop(stop):
    ids = stop.get("ids") or []
    if not ids:
        return ""

    try:
        first_id = int(ids[0])
    except (TypeError, ValueError):
        return ""

    order = RunSheet.objects.filter(id=first_id).only("customer_id").first()
    return order.customer_id if order else ""


def _copy_cell(source, target):
    target.value = source.value
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _shift_block_right_for_code_column(ws, block):
    """
    Converts a generated 9-column transport block:
        Customer Name | City | Weight | Skids | Bundles | Coils | Closes at | Pickup | spacer

    into a 10-column transport block:
        Code | Customer Name | City | Weight | Skids | Bundles | Coils | Closes at | Pickup | spacer

    This keeps the code visible in the Excel file so it moves naturally when rows are copied/pasted.
    """
    start_col = block["start_col"]

    for row_num in range(block["header_row"], block["total_row"] + 1):
        for col in range(start_col + 8, start_col - 1, -1):
            _copy_cell(
                ws.cell(row=row_num, column=col),
                ws.cell(row=row_num, column=col + 1),
            )

    # The code column uses the old customer-name column style.
    code_header = ws.cell(row=block["header_row"], column=start_col)
    code_header.value = "Code"

    for row_num in range(block["start_row"], block["end_row"] + 1):
        ws.cell(row=row_num, column=start_col).value = None


def _clear_spacer_column(ws, block):
    """
    The last column in each region block is only a spacer.
    Clear it after shifting so old IDs or copied values cannot appear there.
    """
    spacer_col = block["start_col"] + 9

    for row_num in range(block["header_row"], block["total_row"] + 1):
        ws.cell(row=row_num, column=spacer_col).value = None


def _remove_legacy_runsheet_metadata(workbook):
    """
    Remove all legacy internal RunSheet metadata from the generated workbook.

    Older export code can add strings like "runsheet_id=206" as hidden values,
    comments, hyperlinks, or data-validation prompts. The visible Code column
    replaces that system, so these values should not be shipped to transport.
    """
    for ws in workbook.worksheets:
        # Clear old hidden metadata columns and any far-right helper area.
        for col_idx in list(range(21, 61)):  # U:BH
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = None
                cell.comment = None
                cell.hyperlink = None

        # Remove any legacy runsheet_id values/comments that may be attached to visible cells.
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "runsheet_id" in cell.value.lower():
                    cell.value = None
                if cell.comment and "runsheet_id" in str(cell.comment.text).lower():
                    cell.comment = None
                if cell.hyperlink and "runsheet_id" in str(cell.hyperlink.target).lower():
                    cell.hyperlink = None

        # Remove data validations that contain legacy runsheet metadata.
        if getattr(ws, "data_validations", None):
            kept_validations = []
            for dv in ws.data_validations.dataValidation:
                dv_text = " ".join(
                    str(value or "")
                    for value in [
                        getattr(dv, "formula1", ""),
                        getattr(dv, "formula2", ""),
                        getattr(dv, "prompt", ""),
                        getattr(dv, "promptTitle", ""),
                        getattr(dv, "error", ""),
                        getattr(dv, "errorTitle", ""),
                    ]
                ).lower()
                if "runsheet_id" not in dv_text:
                    kept_validations.append(dv)

            ws.data_validations.dataValidation = kept_validations
            ws.data_validations.count = len(kept_validations)

        # Ensure old metadata columns remain hidden even after being cleared.
        ws.column_dimensions["AA"].hidden = True
        ws.column_dimensions["AB"].hidden = True


def _autosize_visible_transport_columns(ws):
    """
    Widen A:T so headers and values display instead of being cut off.
    Keeps spacer columns J and T narrow.
    """
    min_widths = {
        "A": 12,   # Code
        "B": 24,   # Customer Name
        "C": 16,   # City
        "D": 12,   # Weight
        "E": 10,   # Skids
        "F": 11,   # Bundles
        "G": 9,    # Coils
        "H": 14,   # Closes at
        "I": 10,   # Pickup
        "J": 4,    # Spacer
        "K": 12,   # Code
        "L": 24,   # Customer Name
        "M": 16,   # City
        "N": 12,   # Weight
        "O": 10,   # Skids
        "P": 11,   # Bundles
        "Q": 9,    # Coils
        "R": 14,   # Closes at
        "S": 10,   # Pickup
        "T": 4,    # Spacer
    }

    max_widths = {
        "B": 45,
        "C": 28,
        "L": 45,
        "M": 28,
    }

    for col_idx in range(1, 21):  # A:T
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        calculated_width = max_length + 2
        min_width = min_widths.get(col_letter, 10)
        max_width = max_widths.get(col_letter, 18)
        ws.column_dimensions[col_letter].width = min(max(calculated_width, min_width), max_width)

    # Force spacer columns narrow and empty-looking.
    ws.column_dimensions["J"].width = 4
    ws.column_dimensions["T"].width = 4


def _apply_readable_alignment(ws):
    for row in ws.iter_rows(min_row=1, max_row=75, min_col=1, max_col=20):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical="center",
                wrap_text=False,
            )


def _write_customer_codes(workbook, shipping_date):
    """
    Calls the existing export first, then adds a visible Code column to each region block.
    Old hidden AA/AB metadata is cleared so stale row IDs cannot affect future imports.
    """
    for ws in workbook.worksheets:
        for region, block in base_views.TRANSPORT_REGION_BLOCKS.items():
            _shift_block_right_for_code_column(ws, block)
            _clear_spacer_column(ws, block)

            metadata_col = _metadata_col_for_block(block)
            if metadata_col:
                for row_num in range(block["start_row"], block["end_row"] + 1):
                    ws.cell(row=row_num, column=metadata_col).value = None

            stops = base_views.build_transport_stops_for_region(region, shipping_date)
            for index, stop in enumerate(stops):
                row_num = block["start_row"] + index
                if row_num > block["end_row"]:
                    break
                ws.cell(row=row_num, column=block["start_col"]).value = _customer_code_for_stop(stop)

            # Rebuild totals because formulas copied by openpyxl do not automatically shift references.
            total_row = block["total_row"]
            start_row = block["start_row"]
            end_row = block["end_row"]
            start_col = block["start_col"]

            weight_col = start_col + 3
            skids_col = start_col + 4
            bundles_col = start_col + 5
            coils_col = start_col + 6

            ws.cell(row=total_row, column=weight_col).value = (
                f"=SUM({ws.cell(start_row, weight_col).coordinate}:{ws.cell(end_row, weight_col).coordinate})"
            )
            ws.cell(row=total_row, column=skids_col).value = (
                f"=SUM({ws.cell(start_row, skids_col).coordinate}:{ws.cell(end_row, skids_col).coordinate})"
            )
            ws.cell(row=total_row, column=bundles_col).value = (
                f"=SUM({ws.cell(start_row, bundles_col).coordinate}:{ws.cell(end_row, bundles_col).coordinate})"
            )
            ws.cell(row=total_row, column=coils_col).value = (
                f"=SUM({ws.cell(start_row, coils_col).coordinate}:{ws.cell(end_row, coils_col).coordinate})"
            )

    _remove_legacy_runsheet_metadata(workbook)

    for ws in workbook.worksheets:
        # Keep old metadata columns hidden/empty for backwards safety, but new matching uses visible Code.
        ws.column_dimensions["AA"].hidden = True
        ws.column_dimensions["AB"].hidden = True
        ws.print_area = "A1:T75"

        _autosize_visible_transport_columns(ws)
        _apply_readable_alignment(ws)


@login_required
def export_run_sheet_excel(request):
    """
    Wrapper around the original Excel export.

    It preserves the existing generated workbook/layout, then adds a visible Code column
    at the start of each transport region block. This avoids hidden-cell mismatch issues
    when the transport sheet is copy/pasted or rearranged.
    """
    shipping_date = base_views.get_selected_shipping_date(request)
    response = base_views.export_run_sheet_excel(request)

    workbook = load_workbook(BytesIO(response.content))
    _write_customer_codes(workbook, shipping_date)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response.content = output.getvalue()
    return response
