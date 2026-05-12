import re
import unicodedata
from difflib import SequenceMatcher
from collections import defaultdict

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import load_workbook

from .models import RunSheet, TransportImportBatch, TransportImportRow, TransportImportPreviousState
from .views import TRANSPORT_REGION_BLOCKS, get_selected_shipping_date, redirect_run_sheet_for_date


LEGAL_WORDS = {
    "INC", "INCORPORATED", "LTD", "LTEE", "LTÉE", "LIMITED", "CORP", "CORPORATION",
    "CO", "COMPANY", "THE", "LES", "LE", "LA", "DES", "DE", "DU", "DISTRIBUTION",
}

TIME_RE = re.compile(r"\b(?:[01]?\d|2[0-3])[:hH][0-5]\d\b|\b(?:[1-9]|1[0-2])\s*(?:AM|PM|A\.M\.|P\.M\.)\b", re.I)
TRANSPORT_PLACEHOLDER_ORDER = "TRANSPORT IMPORT"
TRANSPORT_PLACEHOLDER_PREPARED_BY = "Transport Import"


def clean_match_text(value):
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    parts = [p for p in text.split() if p not in LEGAL_WORDS]
    return " ".join(parts)


def extract_time(value):
    if value is None:
        return ""
    text = str(value).strip()
    match = TIME_RE.search(text)
    return match.group(0).replace("h", ":").replace("H", ":") if match else ""


def clean_driver_name(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = TIME_RE.sub("", text)
    text = re.sub(r"\b(start|time|départ|depart|driver|chauffeur)\b", "", text, flags=re.I)
    text = re.sub(r"[:\-–—]+", " ", text)
    return " ".join(text.split()).strip()


def detect_start_time(ws, block):
    cells_to_check = []
    driver_cell = ws[block["driver_cell"]]
    cells_to_check.append(driver_cell.value)

    # Check a few cells near the region/driver header. Transport may type start time beside or below driver name.
    for row in range(max(1, driver_cell.row - 1), driver_cell.row + 3):
        for col in range(max(1, driver_cell.column - 1), driver_cell.column + 5):
            cells_to_check.append(ws.cell(row=row, column=col).value)

    for value in cells_to_check:
        found = extract_time(value)
        if found:
            return found
    return ""


def score_match(imported_name, imported_city, order):
    name_score = SequenceMatcher(None, clean_match_text(imported_name), clean_match_text(order.customer_name)).ratio()
    city_score = SequenceMatcher(None, clean_match_text(imported_city), clean_match_text(order.city)).ratio()

    if imported_city:
        return round((name_score * 0.8) + (city_score * 0.2), 3)
    return round(name_score, 3)


def stop_group_key(order):
    return (
        order.customer_id or "",
        order.customer_name or "",
        order.city or "",
        order.address or "",
        order.postal_code or "",
        order.closing_time or "",
        bool(order.is_pickup),
        bool(order.is_return),
    )


def build_current_stop_options(shipping_date):
    grouped = {}
    orders = RunSheet.objects.filter(shipping_date=shipping_date).order_by("region", "load_index", "customer_name", "id")

    for order in orders:
        key = stop_group_key(order)
        if key not in grouped:
            grouped[key] = {
                "ids": [],
                "label": f"{order.customer_name or ''} — {order.city or ''} — {order.region or ''}",
                "order_numbers": [],
            }
        grouped[key]["ids"].append(str(order.id))
        if order.order_number:
            grouped[key]["order_numbers"].append(order.order_number)

    options = []
    for item in grouped.values():
        order_refs = " / ".join(item["order_numbers"])
        label = item["label"]
        if order_refs:
            label = f"{label} ({order_refs})"
        options.append({"value": ",".join(item["ids"]), "label": label})

    return sorted(options, key=lambda x: x["label"])


def ids_from_cell(value):
    if value in (None, ""):
        return []
    text = str(value)
    return [int(x) for x in re.findall(r"\d+", text)]


def customer_code_from_cell(value):
    if value in (None, ""):
        return ""
    return str(value).strip()


def get_customer_code_column(block):
    # New exports should use customer_code_col. Old exports still use hidden_ids_col.
    return block.get("customer_code_col") or block.get("hidden_ids_col")


def grouped_ids_for_order(order, candidates):
    key = stop_group_key(order)
    return [o.id for o in candidates if stop_group_key(o) == key]


def row_matches_order_visible_text(row_data, order, minimum_score=0.58):
    imported_name = row_data.get("imported_customer_name") or ""
    imported_city = row_data.get("imported_city") or ""

    # If the row only contains metadata and no visible customer text, allow metadata matching.
    if not imported_name and not imported_city:
        return True, 1.0

    score = score_match(imported_name, imported_city, order)
    return score >= minimum_score, score


def parse_website_export(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    parsed_rows = []

    for ws in wb.worksheets:
        for region, block in TRANSPORT_REGION_BLOCKS.items():
            raw_driver = ws[block["driver_cell"]].value or ""
            driver = clean_driver_name(raw_driver)
            start_time = detect_start_time(ws, block)
            metadata_col = get_customer_code_column(block)
            stop_no = 1

            for row_num in range(block["start_row"], block["end_row"] + 1):
                start_col = block["start_col"]
                customer_name = ws.cell(row=row_num, column=start_col).value
                city = ws.cell(row=row_num, column=start_col + 1).value
                metadata_value = ws.cell(row=row_num, column=metadata_col).value if metadata_col else None
                hidden_ids = ids_from_cell(metadata_value)
                customer_code = customer_code_from_cell(metadata_value)

                if not customer_name and not city and not hidden_ids and not customer_code:
                    continue

                parsed_rows.append({
                    "sheet_name": ws.title,
                    "source_row_number": row_num,
                    "imported_run_name": region,
                    "imported_driver": driver,
                    "imported_truck": "",
                    "imported_start_time": start_time,
                    "imported_stop_number": stop_no,
                    "imported_customer_name": str(customer_name or "").strip(),
                    "imported_city": str(city or "").strip(),
                    "hidden_ids": hidden_ids,
                    "customer_code": customer_code,
                })
                stop_no += 1

    return parsed_rows


def auto_match_row(row_data, shipping_date):
    imported_name = row_data.get("imported_customer_name") or ""
    imported_city = row_data.get("imported_city") or ""
    candidates = RunSheet.objects.filter(shipping_date=shipping_date)

    if not candidates.exists():
        return [], 0, "unmatched"

    customer_code = (row_data.get("customer_code") or "").strip()
    if customer_code:
        code_candidates = candidates.filter(customer_id=customer_code).order_by("load_index", "id")
        if code_candidates.exists():
            best_order = None
            best_score = 0
            for order in code_candidates:
                visible_match, score = row_matches_order_visible_text(row_data, order, minimum_score=0.58)
                if visible_match and score >= best_score:
                    best_order = order
                    best_score = score

            if best_order:
                return grouped_ids_for_order(best_order, candidates), max(best_score, 0.95), "matched"

    # Backwards compatibility for older Excel exports that still contain RunSheet row IDs.
    # Do not blindly trust these IDs: if a hidden ID stayed behind during copy/paste,
    # the visible customer text will not match and we should fall back to fuzzy matching instead.
    hidden_ids = row_data.get("hidden_ids") or []
    if hidden_ids:
        valid_orders = list(candidates.filter(id__in=hidden_ids))
        if valid_orders:
            visible_matches = [
                row_matches_order_visible_text(row_data, order, minimum_score=0.82)[0]
                for order in valid_orders
            ]
            if all(visible_matches):
                return [order.id for order in valid_orders], 1.0, "matched"

    if not imported_name:
        return [], 0, "unmatched"

    best_order = None
    best_score = 0
    for order in candidates:
        score = score_match(imported_name, imported_city, order)
        if score > best_score:
            best_score = score
            best_order = order

    if not best_order:
        return [], 0, "unmatched"

    matched_ids = grouped_ids_for_order(best_order, candidates)
    status = "matched" if best_score >= 0.82 else "review" if best_score >= 0.58 else "unmatched"
    return matched_ids, best_score, status


def create_unmatched_transport_placeholder(row, batch):
    """
    Creates a lightweight RunSheet row for a stop that exists only on the transport sheet.
    It is clearly marked and can be removed by Undo.
    """
    run_name = row.imported_run_name or "Transport Run"
    return RunSheet.objects.create(
        customer_id="",
        shipping_date=batch.shipping_date,
        customer_name=row.imported_customer_name or "Unmatched Transport Stop",
        address="",
        city=row.imported_city or "",
        postal_code="",
        region=run_name,
        order_number=TRANSPORT_PLACEHOLDER_ORDER,
        prepared_by=TRANSPORT_PLACEHOLDER_PREPARED_BY,
        line_items=0,
        closing_time="",
        weight=0,
        skids=0,
        bundles=0,
        coils=0,
        load_index=row.imported_stop_number or row.sort_order,
        driver_name=row.imported_driver or "",
        transport_run_name=run_name,
        transport_driver=row.imported_driver or "",
        transport_truck=row.imported_truck or "",
        transport_start_time=row.imported_start_time or "",
        transport_stop_number=row.imported_stop_number or row.sort_order,
        transport_import_batch=batch,
    )


@login_required
def upload_transport_import(request):
    if request.method != "POST":
        return redirect_run_sheet_for_date(get_selected_shipping_date(request))

    shipping_date = get_selected_shipping_date(request)
    excel_file = request.FILES.get("excel_file")

    if not excel_file:
        messages.error(request, "Please choose an Excel file first.")
        return redirect_run_sheet_for_date(shipping_date)

    batch = TransportImportBatch.objects.create(
        shipping_date=shipping_date,
        original_filename=excel_file.name,
        uploaded_by=request.user.username if request.user.is_authenticated else "",
        status="review",
    )

    try:
        parsed_rows = parse_website_export(excel_file)
    except Exception as exc:
        batch.status = "failed"
        batch.notes = str(exc)
        batch.save(update_fields=["status", "notes"])
        messages.error(request, f"Could not read that Excel file: {exc}")
        return redirect_run_sheet_for_date(shipping_date)

    if not parsed_rows:
        batch.status = "failed"
        batch.notes = "No stops were detected in the uploaded Excel file."
        batch.save(update_fields=["status", "notes"])
        messages.error(request, "No stops were detected in the uploaded Excel file.")
        return redirect_run_sheet_for_date(shipping_date)

    for sort_order, row_data in enumerate(parsed_rows, start=1):
        matched_ids, confidence, status = auto_match_row(row_data, shipping_date)
        TransportImportRow.objects.create(
            batch=batch,
            sort_order=sort_order,
            source_sheet_name=row_data["sheet_name"],
            source_row_number=row_data["source_row_number"],
            imported_run_name=row_data["imported_run_name"],
            imported_driver=row_data["imported_driver"],
            imported_truck=row_data["imported_truck"],
            imported_start_time=row_data["imported_start_time"],
            imported_stop_number=row_data["imported_stop_number"],
            imported_customer_name=row_data["imported_customer_name"],
            imported_city=row_data["imported_city"],
            matched_run_sheet_ids=",".join(str(i) for i in matched_ids),
            confidence=confidence,
            status=status,
        )

    messages.info(request, "Transport sheet uploaded. Review the matches before applying the order.")
    return redirect("review_transport_import", batch_id=batch.id)


@login_required
def review_transport_import(request, batch_id):
    batch = get_object_or_404(TransportImportBatch, pk=batch_id)
    rows = batch.rows.all().order_by("sort_order", "id")
    options = build_current_stop_options(batch.shipping_date)

    if request.method == "POST":
        for row in rows:
            selected_ids = request.POST.get(f"match_{row.id}", "").strip()
            row.matched_run_sheet_ids = selected_ids
            row.status = "matched" if selected_ids else "unmatched"
            row.save(update_fields=["matched_run_sheet_ids", "status"])

        if request.POST.get("apply_now") == "1":
            return redirect("apply_transport_import", batch_id=batch.id)

        messages.success(request, "Matches saved.")
        return redirect("review_transport_import", batch_id=batch.id)

    matched_count = rows.filter(status="matched").count()
    review_count = rows.filter(status="review").count()
    unmatched_count = rows.exclude(status__in=["matched", "review"]).count()

    return render(request, "core/transport_import_review.html", {
        "batch": batch,
        "rows": rows,
        "options": options,
        "matched_count": matched_count,
        "review_count": review_count,
        "unmatched_count": unmatched_count,
    })


@login_required
def apply_transport_import(request, batch_id):
    batch = get_object_or_404(TransportImportBatch, pk=batch_id)
    rows = batch.rows.all().order_by("sort_order", "id")

    if not rows.exists():
        messages.error(request, "There are no rows to apply.")
        return redirect("review_transport_import", batch_id=batch.id)

    with transaction.atomic():
        TransportImportPreviousState.objects.filter(batch=batch).delete()

        # If this batch was applied before, remove its old unmatched placeholder rows before recreating them.
        RunSheet.objects.filter(
            transport_import_batch=batch,
            customer_id="",
            order_number=TRANSPORT_PLACEHOLDER_ORDER,
            prepared_by=TRANSPORT_PLACEHOLDER_PREPARED_BY,
        ).delete()

        touched_ids = set()
        created_unmatched_count = 0

        for row in rows:
            ids = [int(x) for x in row.matched_run_sheet_ids.split(",") if x.strip().isdigit()]

            if not ids:
                create_unmatched_transport_placeholder(row, batch)
                created_unmatched_count += 1
                continue

            for run_item in RunSheet.objects.select_for_update().filter(id__in=ids, shipping_date=batch.shipping_date):
                if run_item.id not in touched_ids:
                    TransportImportPreviousState.objects.create(
                        batch=batch,
                        run_sheet_id=run_item.id,
                        previous_transport_run_name=run_item.transport_run_name,
                        previous_transport_driver=run_item.transport_driver,
                        previous_transport_truck=run_item.transport_truck,
                        previous_transport_start_time=run_item.transport_start_time,
                        previous_transport_stop_number=run_item.transport_stop_number,
                        previous_transport_import_batch_id=run_item.transport_import_batch_id,
                        previous_driver_name=run_item.driver_name,
                        previous_load_index=run_item.load_index,
                    )
                    touched_ids.add(run_item.id)

                run_item.transport_run_name = row.imported_run_name or run_item.region or "Transport Run"
                run_item.transport_driver = row.imported_driver or ""
                run_item.transport_truck = row.imported_truck or ""
                run_item.transport_start_time = row.imported_start_time or ""
                run_item.transport_stop_number = row.imported_stop_number or row.sort_order
                run_item.transport_import_batch = batch
                if row.imported_driver:
                    run_item.driver_name = row.imported_driver
                run_item.load_index = row.imported_stop_number or row.sort_order
                run_item.save(update_fields=[
                    "transport_run_name", "transport_driver", "transport_truck", "transport_start_time",
                    "transport_stop_number", "transport_import_batch", "driver_name", "load_index",
                ])

        batch.status = "applied"
        batch.applied_at = timezone.now()
        batch.save(update_fields=["status", "applied_at"])

    messages.success(
        request,
        f"Transport order applied. {len(touched_ids)} existing rows were updated and {created_unmatched_count} unmatched transport stops were imported as-is. You can undo this import if needed."
    )
    return redirect("transport_run_sheet_view", batch_id=batch.id)


@login_required
def undo_transport_import(request, batch_id):
    batch = get_object_or_404(TransportImportBatch, pk=batch_id)

    if request.method != "POST":
        return redirect("transport_import_history")

    states = TransportImportPreviousState.objects.filter(batch=batch)
    placeholder_count = RunSheet.objects.filter(
        transport_import_batch=batch,
        customer_id="",
        order_number=TRANSPORT_PLACEHOLDER_ORDER,
        prepared_by=TRANSPORT_PLACEHOLDER_PREPARED_BY,
    ).count()

    if not states.exists() and placeholder_count == 0:
        messages.error(request, "No saved previous state or imported unmatched rows were found for this import.")
        return redirect("transport_import_history")

    with transaction.atomic():
        # Remove unmatched rows that were created only from the transport sheet.
        RunSheet.objects.filter(
            transport_import_batch=batch,
            customer_id="",
            order_number=TRANSPORT_PLACEHOLDER_ORDER,
            prepared_by=TRANSPORT_PLACEHOLDER_PREPARED_BY,
        ).delete()

        for state in states:
            RunSheet.objects.filter(id=state.run_sheet_id).update(
                transport_run_name=state.previous_transport_run_name,
                transport_driver=state.previous_transport_driver,
                transport_truck=state.previous_transport_truck,
                transport_start_time=state.previous_transport_start_time,
                transport_stop_number=state.previous_transport_stop_number,
                transport_import_batch_id=state.previous_transport_import_batch_id,
                driver_name=state.previous_driver_name,
                load_index=state.previous_load_index,
            )

        batch.status = "undone"
        batch.undone_at = timezone.now()
        batch.save(update_fields=["status", "undone_at"])

    messages.success(request, "Transport import was undone. Previous order was restored and unmatched imported stops were removed.")
    return redirect_run_sheet_for_date(batch.shipping_date)


@login_required
def transport_import_history(request):
    selected_shipping_date = get_selected_shipping_date(request)
    batches = TransportImportBatch.objects.filter(shipping_date=selected_shipping_date).order_by("-created_at")
    return render(request, "core/transport_import_history.html", {
        "batches": batches,
        "selected_shipping_date": selected_shipping_date,
    })


@login_required
def transport_run_sheet_view(request, batch_id=None):
    selected_shipping_date = get_selected_shipping_date(request)

    if batch_id:
        batch = get_object_or_404(TransportImportBatch, pk=batch_id)
    else:
        batch = TransportImportBatch.objects.filter(
            shipping_date=selected_shipping_date,
            status="applied",
        ).order_by("-applied_at", "-created_at").first()

    if not batch:
        messages.warning(request, "No applied transport import exists for this shipping date yet.")
        return redirect_run_sheet_for_date(selected_shipping_date)

    if request.method == "POST":
        original_run_name = request.POST.get("original_run_name", "").strip()
        original_driver = request.POST.get("original_driver", "").strip()
        original_start_time = request.POST.get("original_start_time", "").strip()
        new_driver = request.POST.get("transport_driver", "").strip()
        new_start_time = request.POST.get("transport_start_time", "").strip()

        run_orders = RunSheet.objects.filter(
            shipping_date=batch.shipping_date,
            transport_import_batch=batch,
            transport_run_name=original_run_name,
            transport_driver=original_driver,
            transport_start_time=original_start_time,
        )

        updated = run_orders.update(
            transport_driver=new_driver,
            transport_start_time=new_start_time,
            driver_name=new_driver,
        )

        messages.success(request, f"Updated driver/start time for {updated} rows.")
        return redirect("transport_run_sheet_view", batch_id=batch.id)

    orders = RunSheet.objects.filter(
        shipping_date=batch.shipping_date,
        transport_import_batch=batch,
    ).order_by("transport_run_name", "transport_driver", "transport_start_time", "transport_stop_number", "customer_name", "order_number", "id")

    grouped = defaultdict(lambda: {
        "orders": [],
        "transport_run_name": "",
        "transport_driver": "",
        "transport_start_time": "",
        "totals": {"weight": 0, "skids": 0, "bundles": 0, "coils": 0},
    })

    for order in orders:
        run_name = order.transport_run_name or order.region or "Transport Run"
        driver = order.transport_driver or ""
        start_time = order.transport_start_time or ""
        group_key = f"{run_name}|{driver}|{start_time}"
        run_label = run_name
        if driver:
            run_label = f"{run_label} — {driver}"
        if start_time:
            run_label = f"{run_label} @ {start_time}"

        grouped[group_key]["label"] = run_label
        grouped[group_key]["transport_run_name"] = run_name
        grouped[group_key]["transport_driver"] = driver
        grouped[group_key]["transport_start_time"] = start_time
        grouped[group_key]["orders"].append(order)
        grouped[group_key]["totals"]["weight"] += order.weight or 0
        grouped[group_key]["totals"]["skids"] += order.skids or 0
        grouped[group_key]["totals"]["bundles"] += order.bundles or 0
        grouped[group_key]["totals"]["coils"] += order.coils or 0

    return render(request, "core/transport_run_sheet.html", {
        "batch": batch,
        "grouped_runs": dict(grouped),
        "selected_shipping_date": batch.shipping_date,
    })
