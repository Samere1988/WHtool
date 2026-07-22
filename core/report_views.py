from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from .employee_choices import WAREHOUSE_EMPLOYEES
import re
from uuid import uuid4
from django.db.models import Count
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import (
    require_GET,
    require_http_methods,
    require_POST,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import (
    CycleCount,
    CycleCountCounter,
    CycleCountItem,
    HymusTransferItem,
    InventoryItem,
    InventoryReport,
)

MAX_REPORT_SIZE = 50 * 1024 * 1024
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
TRANSFER_ROWS_PER_PAGE = 10
SHEET_DESCRIPTION_PREFIXES = (
    "SSH",
    "SPL",
    "APL",
    "ASH",
    "ACP",
    "SCL",
    "ACL",
)

EXCLUDED_CYCLE_COUNT_LOCATIONS = {
    "24",
    "AMERINOX",
    "CFF",
    "COILEX",
    "DANACA",
    "LAPORTE",
    "LEVELTEK",
    "MAINST-ELK",
    "MAINST-HAR",
    "MAINSTEEL",
    "NAS",
    "SCRAPBIN",
    "SECOND",
    "SSG",
}

LONG_PRODUCT_RACKS = {
    letter * 2
    for letter in "ABCDEFGHIJKLMNOPQRSTU"
}

LONG_PRODUCT_RACKS.update({
    "VV",
    "WW",
    "XX",
    "YY",
    "Y",
    "Z",
    "LX",
    "LZ",
})

class ReportUploadError(ValueError):
    pass


def normalize_log_number(value):
    return str(value or "").strip().upper()


def clean_text(value):
    if value in (None, ""):
        return ""

    return " ".join(str(value).split())
def bin_location_sort_key(item):
    location = clean_text(
        item.bin_location
    ).upper()

    return [
        int(part)
        if part.isdigit()
        else part
        for part in re.split(
            r"(\d+)",
            location,
        )
    ]


def is_sheet_material(item):
    description = clean_text(
        item.description
    ).upper()

    return description.startswith(
        SHEET_DESCRIPTION_PREFIXES
    )
def natural_text_sort_key(value):
    return tuple(
        (0, int(part))
        if part.isdigit()
        else (1, part)
        for part in re.split(
            r"(\d+)",
            clean_text(value).upper(),
        )
        if part != ""
    )


def cycle_count_rack_key(bin_location):
    location = clean_text(
        bin_location
    ).upper()

    if (
        not location
        or location
        in EXCLUDED_CYCLE_COUNT_LOCATIONS
    ):
        return None

    compact_location = re.sub(
        r"\s+",
        "",
        location,
    )

    if compact_location == "26HYMUS":
        return "26 HYMUS"

    if location.startswith("BAY"):
        return "BAY"

    # 1A-1, 1A-2, 1A-FLR -> 1A
    match = re.match(
        r"^([1-5][A-Z])(?:-|$)",
        location,
    )

    if match:
        return match.group(1)

    # 26A1, 26A2 -> 26A
    match = re.match(
        r"^(26[A-Z])\d+$",
        location,
    )

    if match:
        return match.group(1)

    # AA1, AA2 and FW5, FW6 -> AA or FW
    match = re.match(
        r"^([A-Z]{2})\d+$",
        location,
    )

    if match:
        return match.group(1)

    # LX-1, SHIP-1, VV-1 -> LX, SHIP or VV
    match = re.match(
        r"^([A-Z]{1,4})-\d+$",
        location,
    )

    if match:
        return match.group(1)

    return location


def cycle_count_category(rack):
    if rack in LONG_PRODUCT_RACKS:
        return CycleCount.Category.LONG_PRODUCTS

    return CycleCount.Category.SHEETS


def add_nullable_quantity(current, incoming):
    if current is None:
        return incoming

    if incoming is None:
        return current

    return current + incoming


def build_cycle_count_groups(report):
    if report is None:
        return []

    groups = {}

    inventory_items = (
        report.items.all()
        .order_by(
            "bin_location",
            "log_number",
            "id",
        )
    )

    for item in inventory_items:
        rack = cycle_count_rack_key(
            item.bin_location
        )

        if rack is None:
            continue

        category = cycle_count_category(
            rack
        )

        group_key = (category, rack)

        if group_key not in groups:
            groups[group_key] = {
                "category": category,
                "rack": rack,
                "items_by_key": {},
            }

        display_bin = clean_text(
            item.bin_location
        ).upper()

        if re.sub(
            r"\s+",
            "",
            display_bin,
        ) == "26HYMUS":
            display_bin = "26 HYMUS"

        item_key = (
            normalize_log_number(
                item.log_number
            ),
            clean_text(item.description),
            display_bin,
        )

        group_item = groups[
            group_key
        ]["items_by_key"].get(
            item_key
        )

        if group_item is None:
            groups[group_key][
                "items_by_key"
            ][item_key] = {
                "log_number": item_key[0],
                "description": item_key[1],
                "bin_location": item_key[2],
                "on_hand_pieces": (
                    item.on_hand_pieces
                ),
                "on_hand_weight": (
                    item.on_hand_weight
                ),
            }
        else:
            # Combine matching physical inventory that was
            # separated by the sales warehouse distinction.
            group_item[
                "on_hand_pieces"
            ] = add_nullable_quantity(
                group_item[
                    "on_hand_pieces"
                ],
                item.on_hand_pieces,
            )

            group_item[
                "on_hand_weight"
            ] = add_nullable_quantity(
                group_item[
                    "on_hand_weight"
                ],
                item.on_hand_weight,
            )

    result = []

    for group in groups.values():
        items = sorted(
            group["items_by_key"].values(),
            key=lambda row: (
                natural_text_sort_key(
                    row["bin_location"]
                ),
                natural_text_sort_key(
                    row["log_number"]
                ),
            ),
        )

        result.append({
            "category": group["category"],
            "rack": group["rack"],
            "token": (
                f"{group['category']}|"
                f"{group['rack']}"
            ),
            "items": items,
            "item_count": len(items),
        })

    result.sort(
        key=lambda group: (
            0
            if group["category"]
            == CycleCount.Category.SHEETS
            else 1,
            natural_text_sort_key(
                group["rack"]
            ),
        )
    )

    return result
def clean_decimal(value, field_name, log_number):
    if value in (None, ""):
        return None

    try:
        number = Decimal(
            str(value).replace(",", "").strip()
        )
    except (InvalidOperation, ValueError) as exc:
        raise ReportUploadError(
            f"Log {log_number}: "
            f"{field_name} is not a valid number."
        ) from exc

    if not number.is_finite():
        raise ReportUploadError(
            f"Log {log_number}: "
            f"{field_name} is not a valid number."
        )

    return number.quantize(Decimal("0.01"))


def clean_integer(value, field_name, log_number):
    number = clean = clean_decimal(
        value,
        field_name,
        log_number,
    )

    if number is None:
        return None

    if number != number.to_integral_value():
        raise ReportUploadError(
            f"Log {log_number}: "
            f"{field_name} must be a whole number."
        )

    return int(number)


def validate_upload(upload):
    extension = Path(upload.name).suffix.lower()

    if extension not in ALLOWED_EXTENSIONS:
        raise ReportUploadError(
            f"{upload.name}: please upload "
            f"an .xlsx or .xlsm file."
        )

    if upload.size > MAX_REPORT_SIZE:
        raise ReportUploadError(
            f"{upload.name}: the file must be "
            f"50 MB or smaller."
        )


def open_report_worksheet(upload):
    validate_upload(upload)

    try:
        upload.seek(0)

        workbook = load_workbook(
            upload,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ReportUploadError(
            f"{upload.name} could not be read "
            f"as an Excel workbook."
        ) from exc

    if "Report" in workbook.sheetnames:
        worksheet = workbook["Report"]
    else:
        worksheet = workbook.active

    return workbook, worksheet


def read_bin_location_report(
    upload,
    expected_warehouse,
):
    """
    Inventory report layout:

    A = Warehouse
    B = Description
    C = Log Number
    D = Location Type (usually F)
    E = Bin Location
    H = Available Weight
    I = Available Pieces
    L = On Hand Weight
    M = On Hand Pieces
    """
    workbook, worksheet = open_report_worksheet(
        upload
    )

    rows = []
    header_found = False

    try:
        for values in worksheet.iter_rows(
            values_only=True
        ):
            values = list(values)

            values.extend(
                [None] * max(
                    0,
                    13 - len(values),
                )
            )

            if (
                normalize_log_number(values[2])
                == "LOG NUMBER"
                and clean_text(values[3]).upper()
                == "BIN LOCATION"
            ):
                header_found = True
                continue

            if not header_found:
                continue

            warehouse = clean_text(values[0])
            log_number = normalize_log_number(
                values[2]
            )

            if (
                warehouse
                != str(expected_warehouse)
                or not log_number
            ):
                continue

            rows.append({
                "warehouse": warehouse,
                "log_number": log_number,
                "description": clean_text(
                    values[1]
                ),
                "bin_location": clean_text(
                    values[4]
                ),
                "available_weight": clean_decimal(
                    values[7],
                    "available weight",
                    log_number,
                ),
                "available_pieces": clean_integer(
                    values[8],
                    "available pieces",
                    log_number,
                ),
                "on_hand_weight": clean_decimal(
                    values[11],
                    "on-hand weight",
                    log_number,
                ),
                "on_hand_pieces": clean_integer(
                    values[12],
                    "on-hand pieces",
                    log_number,
                ),
            })
    finally:
        workbook.close()

    if not header_found:
        raise ReportUploadError(
            f"{upload.name} does not appear to be "
            f"a bin-location report."
        )

    if not rows:
        raise ReportUploadError(
            f"{upload.name} does not contain "
            f"warehouse {expected_warehouse} inventory."
        )

    return rows


def read_remarks_report(upload):
    """
    Remarks report layout:

    A = Warehouse
    E = Log
    X = Remarks 1
    Y = Remarks 2
    Z = Remarks 3
    """
    workbook, worksheet = open_report_worksheet(
        upload
    )

    remarks_by_log = defaultdict(list)
    header_found = False

    try:
        for values in worksheet.iter_rows(
            values_only=True
        ):
            values = list(values)

            values.extend(
                [None] * max(
                    0,
                    26 - len(values),
                )
            )

            if (
                normalize_log_number(values[4])
                == "LOG"
                and clean_text(values[23]).upper()
                == "REMARKS 1"
            ):
                header_found = True
                continue

            if not header_found:
                continue

            warehouse = clean_text(values[0])
            log_number = normalize_log_number(
                values[4]
            )

            if (
                warehouse not in {"20", "21"}
                or not log_number
            ):
                continue

            existing_lowercase = {
                remark.casefold()
                for remark
                in remarks_by_log[log_number]
            }

            for value in values[23:26]:
                remark = clean_text(value)

                if (
                    remark
                    and remark.casefold()
                    not in existing_lowercase
                ):
                    remarks_by_log[
                        log_number
                    ].append(remark)

                    existing_lowercase.add(
                        remark.casefold()
                    )
    finally:
        workbook.close()

    if not header_found:
        raise ReportUploadError(
            f"{upload.name} does not appear to be "
            f"a log-remarks report."
        )

    return remarks_by_log


def combine_reports(
    bin_report_20,
    bin_report_21,
    remarks_report,
):
    warehouse_20_rows = (
        read_bin_location_report(
            bin_report_20,
            expected_warehouse=20,
        )
    )

    warehouse_21_rows = (
        read_bin_location_report(
            bin_report_21,
            expected_warehouse=21,
        )
    )

    remarks_by_log = read_remarks_report(
        remarks_report
    )

    combined_rows = []

    for inventory_row in (
        warehouse_20_rows
        + warehouse_21_rows
    ):
        log_number = inventory_row["log_number"]

        combined_rows.append({
            **inventory_row,
            "remarks": " | ".join(
                remarks_by_log.get(
                    log_number,
                    [],
                )
            ),
        })

    combined_rows.sort(
        key=lambda row: (
            row["log_number"],
            row["warehouse"],
            row["bin_location"],
        )
    )

    return combined_rows


def row_value(row, field_name):
    if isinstance(row, dict):
        return row.get(field_name)

    return getattr(row, field_name)


def build_excel_report(rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory Report"

    headers = [
        "Log Number",
        "Description",
        "Bin Location",
        "Available Pieces",
        "Available Weight",
        "On Hand Pieces",
        "On Hand Weight",
        "Remarks",
    ]

    worksheet.append(headers)

    for row in rows:
        worksheet.append([
            row_value(row, "log_number"),
            row_value(row, "description"),
            row_value(row, "bin_location"),
            row_value(row, "available_pieces"),
            row_value(row, "available_weight"),
            row_value(row, "on_hand_pieces"),
            row_value(row, "on_hand_weight"),
            row_value(row, "remarks"),
        ])

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0D2C54",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            vertical="center"
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )
    worksheet.row_dimensions[1].height = 25

    column_widths = {
        "A": 18,
        "B": 55,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 70,
    }

    for column, width in (
        column_widths.items()
    ):
        worksheet.column_dimensions[
            column
        ].width = width

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet[
            f"D{row_number}"
        ].number_format = "#,##0"

        worksheet[
            f"F{row_number}"
        ].number_format = "#,##0"

        worksheet[
            f"E{row_number}"
        ].number_format = "#,##0.00"

        worksheet[
            f"G{row_number}"
        ].number_format = "#,##0.00"

        worksheet[
            f"B{row_number}"
        ].alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        worksheet[
            f"H{row_number}"
        ].alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    output = BytesIO()

    workbook.save(output)
    workbook.close()

    output.seek(0)

    return output


def replace_saved_inventory(
    rows,
    request,
    uploads,
):
    """
    Replace the current inventory atomically.

    If creating the new inventory fails, the
    previous inventory remains in the database.
    """
    with transaction.atomic():
        InventoryReport.objects.all().delete()

        report = InventoryReport.objects.create(
            uploaded_by=(
                request.user.get_username()
            ),
            warehouse_20_filename=(
                uploads[0].name
            ),
            warehouse_21_filename=(
                uploads[1].name
            ),
            remarks_filename=uploads[2].name,
        )

        InventoryItem.objects.bulk_create(
            [
                InventoryItem(
                    report=report,
                    warehouse=row[
                        "warehouse"
                    ],
                    log_number=row[
                        "log_number"
                    ],
                    description=row[
                        "description"
                    ],
                    bin_location=row[
                        "bin_location"
                    ],
                    available_pieces=row[
                        "available_pieces"
                    ],
                    available_weight=row[
                        "available_weight"
                    ],
                    on_hand_pieces=row[
                        "on_hand_pieces"
                    ],
                    on_hand_weight=row[
                        "on_hand_weight"
                    ],
                    remarks=row["remarks"],
                )
                for row in rows
            ],
            batch_size=1000,
        )

    return report


def current_inventory_report():
    return (
        InventoryReport.objects
        .order_by("-uploaded_at")
        .first()
    )


@login_required
@require_http_methods(["GET", "POST"])
def reports(request):
    error = None

    if request.method == "POST":
        uploads = (
            request.FILES.get(
                "bin_report_20"
            ),
            request.FILES.get(
                "bin_report_21"
            ),
            request.FILES.get(
                "remarks_report"
            ),
        )

        if not all(uploads):
            error = (
                "Choose all three reports before "
                "replacing the inventory."
            )
        else:
            try:
                combined_rows = combine_reports(
                    *uploads
                )

                replace_saved_inventory(
                    combined_rows,
                    request,
                    uploads,
                )

            except ReportUploadError as exc:
                error = str(exc)

            else:
                messages.success(
                    request,
                    (
                        "Inventory replaced "
                        "successfully with "
                        f"{len(combined_rows):,} rows."
                    ),
                )

                return redirect("reports")

    current_report = (
        current_inventory_report()
    )

    if current_report:
        inventory_rows = (
            current_report.items.all()
        )
    else:
        inventory_rows = (
            InventoryItem.objects.none()
        )

    total_rows = inventory_rows.count()

    remarks_count = (
        inventory_rows
        .exclude(remarks="")
        .count()
    )

    return render(
        request,
        "core/reports.html",
        {
            "current_report": current_report,
            "combined_rows": inventory_rows,
            "total_rows": total_rows,
            "remarks_count": remarks_count,
            "error": error,
        },
    )


@login_required
@require_GET
def download_inventory_report(request):
    current_report = (
        current_inventory_report()
    )

    if current_report is None:
        messages.error(
            request,
            (
                "Upload an inventory report "
                "before downloading."
            ),
        )

        return redirect("reports")

    output = build_excel_report(
        current_report.items.all()
    )

    filename = (
        "inventory_report_"
        f"{timezone.localdate().isoformat()}"
        ".xlsx"
    )

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response
@login_required
@require_http_methods(["GET", "POST"])
def cycle_counts(request):
    current_report = (
        current_inventory_report()
    )

    rack_groups = build_cycle_count_groups(
        current_report
    )

    if request.method == "POST":
        if current_report is None:
            messages.error(
                request,
                (
                    "Upload the Inventory Report "
                    "before creating a cycle count."
                ),
            )
            return redirect("cycle_counts")

        selected_tokens = set(
            request.POST.getlist("racks")
        )

        selected_groups = [
            group
            for group in rack_groups
            if group["token"]
            in selected_tokens
        ]

        if not selected_groups:
            messages.error(
                request,
                (
                    "Select at least one rack "
                    "before creating the count."
                ),
            )
            return redirect("cycle_counts")

        batch_id = uuid4()
        username = (
            request.user.get_username()
        )

        with transaction.atomic():
            for group in selected_groups:
                cycle_count = (
                    CycleCount.objects.create(
                        batch_id=batch_id,
                        category=group[
                            "category"
                        ],
                        rack=group["rack"],
                        source_inventory_uploaded_at=(
                            current_report.uploaded_at
                        ),
                        created_by=username,
                    )
                )

                CycleCountItem.objects.bulk_create(
                    [
                        CycleCountItem(
                            cycle_count=cycle_count,
                            position=position,
                            log_number=row[
                                "log_number"
                            ],
                            description=row[
                                "description"
                            ],
                            bin_location=row[
                                "bin_location"
                            ],
                            on_hand_pieces=row[
                                "on_hand_pieces"
                            ],
                            on_hand_weight=row[
                                "on_hand_weight"
                            ],
                        )
                        for position, row
                        in enumerate(
                            group["items"],
                            start=1,
                        )
                    ],
                    batch_size=1000,
                )

        messages.success(
            request,
            (
                f"Created {len(selected_groups)} "
                "cycle-count sheet"
                f"{'s' if len(selected_groups) != 1 else ''}."
            ),
        )

        return redirect(
            "cycle_count_batch_print",
            batch_id=batch_id,
        )

    latest_completed = {}

    completed_counts = (
        CycleCount.objects
        .filter(
            completed_at__isnull=False
        )
        .only(
            "category",
            "rack",
            "completed_at",
        )
        .order_by("-completed_at")
    )

    for count in completed_counts:
        key = (
            count.category,
            count.rack,
        )

        if key not in latest_completed:
            latest_completed[key] = (
                count.completed_at
            )

    open_count_by_rack = defaultdict(int)

    for category, rack in (
        CycleCount.objects
        .filter(completed_at__isnull=True)
        .values_list("category", "rack")
    ):
        open_count_by_rack[
            (category, rack)
        ] += 1

    for group in rack_groups:
        key = (
            group["category"],
            group["rack"],
        )

        group["last_completed_at"] = (
            latest_completed.get(key)
        )

        group["open_count"] = (
            open_count_by_rack.get(
                key,
                0,
            )
        )

    sheet_racks = [
        group
        for group in rack_groups
        if group["category"]
        == CycleCount.Category.SHEETS
    ]

    long_product_racks = [
        group
        for group in rack_groups
        if group["category"]
        == CycleCount.Category.LONG_PRODUCTS
    ]

    history = (
        CycleCount.objects
        .annotate(item_total=Count("items"))
        .prefetch_related("counters")
        .order_by("-created_at", "rack")
    )[:250]

    return render(
        request,
        "core/cycle_counts.html",
        {
            "current_report": current_report,
            "sheet_racks": sheet_racks,
            "long_product_racks": long_product_racks,
            "history": history,
            "employees": WAREHOUSE_EMPLOYEES,
        },
    )


def cycle_counts_for_print(queryset):
    counts = list(
        queryset.prefetch_related("items")
    )

    counts.sort(
        key=lambda count: (
            0
            if count.category
            == CycleCount.Category.SHEETS
            else 1,
            natural_text_sort_key(
                count.rack
            ),
        )
    )

    return counts


@login_required
@require_GET
def cycle_count_batch_print(
    request,
    batch_id,
):
    batch_counts = (
        CycleCount.objects.filter(
            batch_id=batch_id
        )
    )

    counts = cycle_counts_for_print(
        batch_counts
    )

    if not counts:
        messages.error(
            request,
            "That cycle-count batch was not found.",
        )

        return redirect("cycle_counts")

    return render(
        request,
        "core/cycle_count_print.html",
        {
            "counts": counts,
            "is_batch": True,
        },
    )


@login_required
@require_GET
def cycle_count_print(
    request,
    count_id,
):
    cycle_count = get_object_or_404(
        CycleCount,
        pk=count_id,
    )

    counts = cycle_counts_for_print(
        CycleCount.objects.filter(
            pk=cycle_count.pk
        )
    )

    return render(
        request,
        "core/cycle_count_print.html",
        {
            "counts": counts,
            "is_batch": False,
        },
    )


@login_required
@require_POST
def complete_cycle_count(
    request,
    count_id,
):
    selected_names = set(
        request.POST.getlist("counters")
    )

    counter_names = [
        name
        for name in WAREHOUSE_EMPLOYEES
        if name in selected_names
    ]

    if not counter_names:
        messages.error(
            request,
            (
                "Select at least one employee "
                "who completed the count."
            ),
        )
        return redirect("cycle_counts")

    with transaction.atomic():
        cycle_count = get_object_or_404(
            CycleCount.objects.select_for_update(),
            pk=count_id,
        )

        if cycle_count.completed_at:
            messages.info(
                request,
                (
                    f"Rack {cycle_count.rack} "
                    "is already marked completed."
                ),
            )
            return redirect("cycle_counts")

        CycleCountCounter.objects.bulk_create(
            [
                CycleCountCounter(
                    cycle_count=cycle_count,
                    employee_name=name,
                )
                for name in counter_names
            ]
        )

        cycle_count.completed_at = timezone.now()

        # This remains the website user who recorded
        # the completion, not necessarily the workers.
        cycle_count.completed_by = (
            request.user.get_username()
        )

        cycle_count.save(
            update_fields=[
                "completed_at",
                "completed_by",
            ]
        )

    messages.success(
        request,
        (
            f"Rack {cycle_count.rack} was marked "
            f"completed by {', '.join(counter_names)}."
        ),
    )

    return redirect("cycle_counts")

@login_required
@require_POST
def delete_cycle_count(
    request,
    count_id,
):
    cycle_count = get_object_or_404(
        CycleCount,
        pk=count_id,
    )

    if cycle_count.completed_at:
        messages.error(
            request,
            (
                "Completed cycle counts cannot "
                "be deleted."
            ),
        )

        return redirect("cycle_counts")

    rack = cycle_count.rack
    cycle_count.delete()

    messages.warning(
        request,
        (
            f"The open cycle count for rack "
            f"{rack} was deleted."
        ),
    )

    return redirect("cycle_counts")



@login_required
@require_http_methods(["GET", "POST"])
def hymus_transfer(request):
    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "add_inventory":
            current_report = (
                current_inventory_report()
            )

            if current_report is None:
                messages.error(
                    request,
                    (
                        "Upload the Inventory Report "
                        "before selecting a log."
                    ),
                )

                return redirect(
                    "hymus_transfer"
                )

            inventory_item = get_object_or_404(
                InventoryItem,
                pk=request.POST.get(
                    "inventory_item_id"
                ),
                report=current_report,
            )

            _, created = (
                HymusTransferItem.objects
                .get_or_create(
                    warehouse=(
                        inventory_item.warehouse
                    ),
                    log_number=(
                        inventory_item.log_number
                    ),
                    bin_location=(
                        inventory_item.bin_location
                    ),
                    defaults={
                        "description": (
                            inventory_item.description
                        ),
                        "added_by": (
                            request.user
                            .get_username()
                        ),
                    },
                )
            )

            if created:
                messages.success(
                    request,
                    (
                        "Added log "
                        f"{inventory_item.log_number} "
                        "to the transfer."
                    ),
                )
            else:
                messages.info(
                    request,
                    (
                        "That inventory row is "
                        "already on the transfer."
                    ),
                )

        elif action == "add_manual":
            log_number = normalize_log_number(
                request.POST.get(
                    "manual_log_number",
                    "",
                )
            )

            description = clean_text(
                request.POST.get(
                    "manual_description",
                    "",
                )
            )

            bin_location = clean_text(
                request.POST.get(
                    "manual_bin_location",
                    "",
                )
            )

            if not all([
                log_number,
                description,
                bin_location,
            ]):
                messages.error(
                    request,
                    (
                        "Enter the log number, "
                        "description, and location."
                    ),
                )

                return redirect(
                    "hymus_transfer"
                )

            current_report = (
                current_inventory_report()
            )

            if (
                current_report
                and current_report.items.filter(
                    log_number__iexact=(
                        log_number
                    )
                ).exists()
            ):
                messages.warning(
                    request,
                    (
                        f"Log {log_number} exists "
                        "in inventory. Select it "
                        "from the search results "
                        "instead."
                    ),
                )

                return redirect(
                    "hymus_transfer"
                )

            _, created = (
                HymusTransferItem.objects
                .get_or_create(
                    warehouse="",
                    log_number=log_number,
                    bin_location=bin_location,
                    defaults={
                        "description": description,
                        "added_by": (
                            request.user
                            .get_username()
                        ),
                    },
                )
            )

            if created:
                messages.success(
                    request,
                    (
                        f"Added manual log "
                        f"{log_number} to the "
                        "transfer."
                    ),
                )
            else:
                messages.info(
                    request,
                    (
                        "That manual log and "
                        "location are already "
                        "on the transfer."
                    ),
                )

        else:
            messages.error(
                request,
                (
                    "Choose a log before "
                    "adding it."
                ),
            )

        return redirect("hymus_transfer")

    transfer_items = sorted(
        HymusTransferItem.objects.all(),
        key=bin_location_sort_key,
    )

    sheet_items = [
        item
        for item in transfer_items
        if is_sheet_material(item)
    ]

    bundle_items = [
        item
        for item in transfer_items
        if not is_sheet_material(item)
    ]
    sheet_pages = [
        sheet_items[index:index + TRANSFER_ROWS_PER_PAGE]
        for index in range(
            0,
            len(sheet_items),
            TRANSFER_ROWS_PER_PAGE,
        )
    ]

    bundle_pages = [
        bundle_items[index:index + TRANSFER_ROWS_PER_PAGE]
        for index in range(
            0,
            len(bundle_items),
            TRANSFER_ROWS_PER_PAGE,
        )
    ]
    transfer_documents = [
        {
            "key": "sheets",
            "title": "26 Hymus Sheets",
            "items": sheet_items,
            "count": len(sheet_items),
            "pages": sheet_pages,
            "page_count": len(sheet_pages),
        },
        {
            "key": "bundles",
            "title": "26 Hymus Bundles",
            "items": bundle_items,
            "count": len(bundle_items),
            "pages": bundle_pages,
            "page_count": len(bundle_pages),
        },
    ]

    return render(
        request,
        "core/hymus_transfer.html",
        {
            "transfer_items": transfer_items,
            "transfer_count": len(
                transfer_items
            ),
            "sheet_items": sheet_items,
            "sheet_count": len(
                sheet_items
            ),
            "bundle_items": bundle_items,
            "bundle_count": len(
                bundle_items
            ),
            "transfer_documents": (
                transfer_documents
            ),
            "current_report": (
                current_inventory_report()
            ),
            "printed_on": (
                timezone.localtime()
            ),
        },
    )


@login_required
@require_GET
def inventory_log_search(request):
    query = normalize_log_number(
        request.GET.get("q", "")
    )

    current_report = (
        current_inventory_report()
    )

    if (
        not query
        or current_report is None
    ):
        return JsonResponse({
            "results": [],
        })

    inventory_items = list(
        current_report.items
        .filter(
            log_number__icontains=query
        )
        .order_by(
            "log_number",
            "warehouse",
            "bin_location",
        )[:25]
    )

    existing_keys = set(
        HymusTransferItem.objects
        .values_list(
            "warehouse",
            "log_number",
            "bin_location",
        )
    )

    results = [
        {
            "id": item.id,
            "log_number": (
                item.log_number
            ),
            "description": (
                item.description
            ),
            "bin_location": (
                item.bin_location
            ),
            "warehouse": (
                item.warehouse
            ),
            "already_added": (
                item.warehouse,
                item.log_number,
                item.bin_location,
            ) in existing_keys,
        }
        for item in inventory_items
    ]

    return JsonResponse({
        "results": results,
    })

@login_required
@require_POST
def update_hymus_transfer_note(
    request,
    item_id,
):
    item = get_object_or_404(
        HymusTransferItem,
        pk=item_id,
    )

    item.notes = request.POST.get(
        "notes",
        "",
    ).strip()

    item.save(
        update_fields=[
            "notes",
            "updated_at",
        ]
    )

    messages.success(
        request,
        (
            f"Notes saved for log "
            f"{item.log_number}."
        ),
    )

    return redirect("hymus_transfer")


@login_required
@require_POST
def remove_hymus_transfer_item(
    request,
    item_id,
):
    item = get_object_or_404(
        HymusTransferItem,
        pk=item_id,
    )

    log_number = item.log_number

    item.delete()

    messages.success(
        request,
        (
            f"Removed log {log_number} "
            f"from the transfer."
        ),
    )

    return redirect("hymus_transfer")


@login_required
@require_POST
def clear_hymus_transfer(request):
    if (
        request.POST.get("confirm_clear")
        != "yes"
    ):
        messages.error(
            request,
            "The transfer was not cleared.",
        )

        return redirect("hymus_transfer")

    deleted_count, _ = (
        HymusTransferItem.objects
        .all()
        .delete()
    )

    messages.success(
        request,
        (
            "26 Hymus Transfer cleared "
            f"({deleted_count:,} rows removed)."
        ),
    )

    return redirect("hymus_transfer")


@login_required
@require_GET
def cycle_count_stats(request):
    today = timezone.localdate()

    selected_month = request.GET.get(
        "month",
        today.strftime("%Y-%m"),
    )

    try:
        year, month = (
            int(value)
            for value in selected_month.split("-", 1)
        )

        month_start = date(
            year,
            month,
            1,
        )

    except (TypeError, ValueError):
        month_start = today.replace(day=1)
        selected_month = month_start.strftime("%Y-%m")

    if month_start.month == 12:
        next_month = date(
            month_start.year + 1,
            1,
            1,
        )
    else:
        next_month = date(
            month_start.year,
            month_start.month + 1,
            1,
        )

    option_dates = set(
        CycleCount.objects
        .filter(completed_at__isnull=False)
        .dates(
            "completed_at",
            "month",
            order="DESC",
        )
    )

    option_dates.add(
        today.replace(day=1)
    )

    month_options = [
        {
            "value": option.strftime("%Y-%m"),
            "label": option.strftime("%B %Y"),
        }
        for option in sorted(
            option_dates,
            reverse=True,
        )
    ]

    counts = list(
        CycleCount.objects
        .filter(
            completed_at__date__gte=month_start,
            completed_at__date__lt=next_month,
        )
        .annotate(
            item_total=Count("items")
        )
        .prefetch_related("counters")
        .order_by(
            "-completed_at",
            "rack",
        )
    )

    employee_totals = defaultdict(
        lambda: {
            "count_sheets": 0,
            "inventory_lines": 0,
            "racks": set(),
        }
    )

    for cycle_count in counts:
        cycle_count.counter_names = [
            counter.employee_name
            for counter
            in cycle_count.counters.all()
        ]

        for employee_name in cycle_count.counter_names:
            totals = employee_totals[
                employee_name
            ]

            totals["count_sheets"] += 1

            totals["inventory_lines"] += (
                cycle_count.item_total
            )

            totals["racks"].add(
                cycle_count.rack
            )

    employee_rows = []

    for employee_name, totals in employee_totals.items():
        employee_rows.append({
            "employee_name": employee_name,
            "count_sheets": totals[
                "count_sheets"
            ],
            "inventory_lines": totals[
                "inventory_lines"
            ],
            "racks": sorted(
                totals["racks"],
                key=natural_text_sort_key,
            ),
        })

    employee_rows.sort(
        key=lambda row: (
            -row["count_sheets"],
            row["employee_name"],
        )
    )

    return render(
        request,
        "core/stats/cycle_count_stats.html",
        {
            "selected_month": selected_month,
            "month_label": month_start.strftime(
                "%B %Y"
            ),
            "month_options": month_options,
            "counts": counts,
            "employee_rows": employee_rows,
            "total_counts": len(counts),
            "total_employees": len(
                employee_rows
            ),
            "total_inventory_lines": sum(
                count.item_total
                for count in counts
            ),
        },
    )

@login_required
@require_POST
def delete_cycle_count(request, count_id):
    cycle_count = get_object_or_404(
        CycleCount,
        pk=count_id,
    )

    if cycle_count.completed_at:
        messages.error(
            request,
            "Completed cycle counts cannot be deleted.",
        )
        return redirect("cycle_counts")

    rack = cycle_count.rack
    cycle_count.delete()

    messages.success(
        request,
        f"Open cycle count for rack {rack} was deleted.",
    )
    return redirect("cycle_counts")